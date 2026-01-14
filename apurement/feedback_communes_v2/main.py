import csv
import math
import os
import pandas as pd
import re
import unicodedata
import urllib.request
import smtplib
import subprocess
import sys
import zipfile
from datetime import datetime
from dotenv import load_dotenv
from email.message import EmailMessage
from email.mime.application import MIMEApplication
from email.utils import formatdate
from jinja2 import Environment, FileSystemLoader
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter, range_boundaries
from openpyxl.worksheet.table import Table, TableStyleInfo
from pathlib import Path


load_dotenv(override=True)


def downloadFiles(url: str, zipPath: str, fileName: str):
    """
    :param url: URL du ZIP à télécharger
    :param path: chemin local où enregistrer le ZIP
    :param fileName: nom du json dans le zip
    :return: extractPath path of zip extraction
    """

    extractPath = Path(str(zipPath).replace(".zip", ""))

    # --- 1. Download and save ZIP on the disk ---
    print("Téléchargement du ZIP...")
    urllib.request.urlretrieve(url, zipPath)
    print(f"ZIP enregistré dans : {zipPath}")

    # --- 2. Extract ZIP ---
    print("Extraction du ZIP...")
    with zipfile.ZipFile(zipPath, "r") as z:
        z.extractall(extractPath)

        # Throw error if filname not in ZIP content
        if fileName not in z.namelist():
            raise FileNotFoundError(f"Fichier JSON '{fileName}' non trouvé dans le ZIP.")

    # --- 3. Supprimer le ZIP ---
    print("Suppression du ZIP...")
    Path.unlink(zipPath)

    return extractPath


def getDataFrame(path: str, json_fileName: str, excel_fileName: str, filters=[]):
    """
    :param path: chemin local où enregistrer le ZIP
    :param fileName: nom du json dans le zip
    :param map_url: URL de la map pour afficher le point
    :return: df data frame pandas
    """

    # --- 1. Lire le JSON ---
    print("read JSON")
    jsonPath = Path(path) / json_fileName
    df = pd.read_json(jsonPath)

    # filter json content by canton
    print("filter data")
    df = df[df["gdekt"] == "NE"]

    # filter egid in whitelist
    if "whitelist" in filters:
        print("filter whitelist")
        # open and get actual data from whitelist
        whitelist_path = r"C:\dev\regbl_toolbox\whitelist\egid_whitelist_controls.xlsx"
        df_whitelist = pd.read_excel(whitelist_path)
        df_whitelist = df_whitelist[df_whitelist["Date sortie"].isna()]
        # filter df with data from df_whitelist
        df = df[~df["egid"].isin(df_whitelist["EGID"])]

    # add new columns
    print("add new columns")
    df["bat_proj"] = ""
    df["etat_mo"] = ""
    df["Type erreur"] = ""
    df["Aide du SGRF"] = ""

    # sort content
    print("sort data")
    df = df.sort_values(by=["gdenr", "issue_bdg_def", "egid"])

    # save to excel file
    print("export data to excel")

    filepath = path / excel_fileName
    df.to_excel(filepath, index=False)

    return filepath


def _getExcelTableColumnIndex(ws, rowIndex: int, searchText: str, searchLimit=1e3):
    """
    :param ws: Excel Worksheet
    :param rowIndex: fixed row index
    :param searchText: text to find
    :param searchLimit: limit of iteration
    :return: colIndex
    """
    rowIndex = 1
    colIndex = 1
    val = ws.cell(rowIndex, colIndex).value
    while val and searchText and val.lower() != searchText.lower() and colIndex < searchLimit:
        colIndex += 1
        val = ws.cell(rowIndex, colIndex).value

    if colIndex == searchLimit and not val.lower() == searchText.lower():
        colIndex = -1

    return colIndex


def _hintErrorTreatment(textError):
    # fixed params
    default_hint = os.environ["FEEDBACK_COMMUNES_v2_DEFAULT_HINT"]

    hint = []
    textError = textError or ""

    # get all error codes
    error_num_list = [int(x) for x in re.findall(r"\b\d+(?=:)", textError)]

    for err_num in error_num_list:
        match err_num:
            case 1:
                help = "Annoncer le cas au SGRF."
            case 12:
                help = "Contrôler et corriger les coordonnées de l'EGID."
            case 13:
                help = "Contrôler et corriger les coordonnées de l'EGID."
            case 14:
                help = "Contrôler et corriger les coordonnées et le statut (GSTAT) de l'EGID."
            case 21:
                help = "L'EGID n'existe que dans la MO. Faut-il le créer dans le RegBL?"
            case 22:
                help = "Créer un EGID."
            case 23:
                help = "Créer un EGID."
            case 24:
                help = "Créer un EGID."
            case 31:
                help = "Faut-il cadastrer un bâtiment, ou corriger les coordonnées de l'EGID? Ou le bâtiment a-t-il été démoli?"
            case 32:
                help = "Ajouter les coordonnées de l'EGID."
            case 33:
                help = "Faut-il cadastrer un bâtiment, ou corriger les coordonnées et/ou le statut (GSTAT) de l'EGID?"
            case 34:
                help = "Données de la MO inexistantes."
            case 35:
                help = "Réunir les EGID."
            case 41:
                help = "Statut du bâtiment (GSTAT) à corriger."
            case 42:
                help = "Catégorie du bâtiment (GKAT) à corriger."
            case 43:
                help = "Catégorie du bâtiment (GKAT) à corriger (provisoire)."
            case 51:
                help = "L'EGID est lié à plusieurs objets de la MO. Faut-il diviser le bâtiment de la MO (avertir le SGRF si nécessaire)?"
            case 52:
                help = "Les EGID de la MO et du RegBL ne correspondent pas (intersection spatiale). Contacter le SGRF"
            case 61:
                help = "Plusieurs objets de la MO se superposent sour les coordonnées de l'EGID."
            case 62:
                help = "Plusieurs EGID contenus dans une empreinte de la MO. Réunir les EGID, diviser le bâtiment ou supprimer un EGID?"
            case _:
                help = default_hint

        hint.append(f"{help}")
        # hint.append(f"{err_num} - {help}")

    return "\n".join(hint)


def _createTable(ws):
    # get sheet dimensions
    min_col, min_row, max_col, max_row = range_boundaries(ws.dimensions)

    # Create Table object in Excel
    start = f"{get_column_letter(min_col)}{min_row}"
    end = f"{get_column_letter(max_col)}{max_row}"
    table = Table(displayName="ListeErreursNE", ref=f"{start}:{end}")
    style = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
    table.tableStyleInfo = style
    ws.add_table(table)
    return ws


def _createHyperlink(cell, hyperlink):
    if cell.value is None:
        cell.value = "SITN"
    cell.hyperlink = hyperlink
    cell.style = "Hyperlink"


def _createEGIDHyperlink(ws):
    # Fixed data
    map_url = os.environ["FEEDBACK_COMMUNES_v2_MAP_URL"]
    map_zoom = os.environ["FEEDBACK_COMMUNES_v2_MAP_ZOOM"]

    # get sheet dimensions
    _, min_row, max_col, max_row = range_boundaries(ws.dimensions)

    # Get column indices
    col_EGID = _getExcelTableColumnIndex(ws, min_row, "egid", max_col)
    col_coordE = _getExcelTableColumnIndex(ws, min_row, "gkode", max_col)
    col_coordN = _getExcelTableColumnIndex(ws, min_row, "gkodn", max_col)

    for row_i in range(min_row + 1, max_row + 1):
        cell = ws.cell(row_i, col_EGID)
        hyperlink = map_url.format(x=ws.cell(row_i, col_coordE).value, y=ws.cell(row_i, col_coordN).value, map_zoom=map_zoom)
        _createHyperlink(cell, hyperlink)


def excelFineTuning(filepath: str):
    """
    :param filepath: Chemin d'accès au fichier Excel
    """
    # fixed params
    default_hint = os.environ["FEEDBACK_COMMUNES_v2_DEFAULT_HINT"]
    hint_out_lpure = os.environ["FEEDBACK_COMMUNES_v2_HINT_OUT_LPURE"]

    # Load file and sheet
    wb = load_workbook(filepath)
    ws = wb.active

    # get sheet dimensions
    _, min_row, max_col, max_row = range_boundaries(ws.dimensions)

    # Get column
    col_GDENR = _getExcelTableColumnIndex(ws, min_row, "gdenr", max_col)  # colonne témoin
    col_issue_bdg_def = _getExcelTableColumnIndex(ws, min_row, "issue_bdg_def", max_col)
    col_issue_bdg_cat = _getExcelTableColumnIndex(ws, min_row, "issue_bdg_cat", max_col)
    col_issue_issue_22 = _getExcelTableColumnIndex(ws, min_row, "issue_issue_22", max_col)
    col_issue_plz = _getExcelTableColumnIndex(ws, min_row, "issue_plz", max_col)
    col_issue_gde = _getExcelTableColumnIndex(ws, min_row, "issue_gde", max_col)
    col_out_gkode_n = _getExcelTableColumnIndex(ws, min_row, "out_gkode_n", max_col)
    col_out_adr_com = _getExcelTableColumnIndex(ws, min_row, "out_adr_com", max_col)
    col_out_adr_inc = _getExcelTableColumnIndex(ws, min_row, "out_adr_inc", max_col)
    col_out_deinr = _getExcelTableColumnIndex(ws, min_row, "out_deinr", max_col)
    col_out_lpure = _getExcelTableColumnIndex(ws, min_row, "out_lpure", max_col)
    col_hint = _getExcelTableColumnIndex(ws, min_row, "Aide du SGRF", max_col)

    row_i = 2
    while row_i <= max_row and ws.cell(row_i, col_GDENR).value != None:
        hint = []
        # based on text issues
        hint.append(_hintErrorTreatment(ws.cell(row_i, col_issue_bdg_def).value)) if ws.cell(row_i, col_issue_bdg_def).value is not None else None
        hint.append(_hintErrorTreatment(ws.cell(row_i, col_issue_bdg_cat).value)) if ws.cell(row_i, col_issue_bdg_cat).value is not None else None
        hint.append(_hintErrorTreatment(ws.cell(row_i, col_issue_issue_22).value)) if ws.cell(row_i, col_issue_issue_22).value is not None else None
        hint.append(_hintErrorTreatment(ws.cell(row_i, col_issue_plz).value)) if ws.cell(row_i, col_issue_plz).value is not None else None
        hint.append(_hintErrorTreatment(ws.cell(row_i, col_issue_gde).value)) if ws.cell(row_i, col_issue_gde).value is not None else None
        # based on error flags
        hint.append("Corriger ou compléter les coordonnées de l'EGID.") if ws.cell(row_i, col_out_gkode_n).value == 1 else None
        hint.append("Doublon d'adresse: Corriger l'adresse en conséquence.") if ws.cell(row_i, col_out_adr_com).value == 1 else None
        hint.append("Adresse non univoque: Ajouter un numéro de bâtiment (DEINR).") if ws.cell(row_i, col_out_adr_inc).value == 1 else None
        hint.append("Le numéro de maison ne répond pas aux recommandations concernant l'adressage de l'OFS (document 1829-1800).") if ws.cell(row_i, col_out_deinr).value == 1 else None
        # if none above but flag out_lpure
        if len(hint) == 0:
            hint.append(hint_out_lpure) if ws.cell(row_i, col_out_lpure).value == 1 else None
        # default value if really none of the above
        if len(hint) == 0:
            hint.append(default_hint) if ws.cell(row_i, col_out_lpure).value == 1 else None

        ws.cell(row_i, col_hint).value = "\n".join(hint)
        ws.cell(row_i, col_hint).alignment = Alignment(wrap_text=True)

        row_i += 1

    ## 2 remove
    # Create Table object in Excel
    # ws = _createTable(ws)

    filepath = filepath.parent / (filepath.stem + "_filtered" + filepath.suffix)
    wb.save(filepath)
    return filepath


def callFME(filepath):
    # fixed parameters
    path_fme_filter_bat_proj = Path(os.environ["FEEDBACK_COMMUNES_v2_FME_FILTER_BAT_PROJ_PATH"])
    fme_filter_bat_proj_args = os.environ["FEEDBACK_COMMUNES_v2_FME_FILTER_BAT_PROJ_ARGUMENTS"]

    print(f"running '{path_fme_filter_bat_proj.name}'...")
    print(">>>", "fme.exe", path_fme_filter_bat_proj, fme_filter_bat_proj_args.replace("{filepath}", str(filepath)))

    subprocess.run(["fme.exe", path_fme_filter_bat_proj] + fme_filter_bat_proj_args.replace("{filepath}", str(filepath)).split(", "))

    # fixed parameters
    path_fme_issue22 = Path(os.environ["FEEDBACK_COMMUNES_v2_FME_ISSUE22_PATH"])
    fme_issue22_args = os.environ["FEEDBACK_COMMUNES_v2_FME_ISSUE22_ARGUMENTS"]
    print(f"running '{path_fme_issue22.name}'...")
    subprocess.run(["fme.exe", path_fme_issue22] + fme_issue22_args.replace("{filepath}", str(filepath)).split(", "))

    # create Excel table object
    wb = load_workbook(filepath)
    ws = wb.active
    ws = _createTable(ws)
    wb.save(filepath)


def remove_cols_by_names_pandas(filepath, col_names_to_keep):
    # print("Removing cols")
    df = pd.read_excel(filepath)
    # df = df.drop(col_names, axis=1)
    for df_col_name in df.columns.tolist():
        if df_col_name not in col_names_to_keep:
            df = df.drop(df_col_name, axis=1)
            # del df[df_col_name]
    df.to_excel(filepath, index=False)


def remove_cols_by_names_openpyxl(filepath, col_names, titleRowIdx=1):
    # print("Removing cols")
    wb = load_workbook(filepath)
    ws = wb.active

    # add egid link
    _createEGIDHyperlink(ws)

    # get sheet dimensions
    _, min_row, max_col, _ = range_boundaries(ws.dimensions)

    # search and remove columnes
    for col_name in col_names:
        titleColIdx = min_row
        while titleColIdx <= max_col:
            if ws.cell(titleRowIdx, titleColIdx).value.lower() == col_name.lower():
                break
            titleColIdx += 1

        if ws.cell(titleRowIdx, titleColIdx).value.lower() == col_name.lower():
            # delete col
            ws.delete_cols(titleColIdx, 1)
            max_col -= 1

    ws = _createTable(ws)
    _auto_resize_worksheet(ws)

    wb.save(filepath)


def _string_normalizer(s):
    # removes accents from string
    s = unicodedata.normalize("NFKD", s)
    return "".join([c for c in s if not unicodedata.combining(c)])


def _string_special_characters(s):
    # removes special characters from string
    s_arr = []
    for char in s:
        s_arr.append(char if char.isalnum() else "_")
    return "".join(s_arr)


def _excel_municipalityFeedback_filename_setter(commune_id, commune_name, dt, suffix, extension):
    commune_name = _string_normalizer(commune_name)
    commune_name = _string_special_characters(commune_name)
    return f"{commune_id}_{commune_name}_{suffix}_{dt}.{extension}"


def exportMunicipalityFile(filepath, municipality_filepath, date, filters=[], attributes_to_remove=[]):
    df = pd.read_excel(filepath)
    df["Type erreur"] = df["Type erreur"].astype("string")

    municipalities_registry = get_municipalities_registry(municipality_filepath, canton="NE")
    unique_gdenr_list = df["gdenr"].unique()

    with open(filepath.parent / f"{date}-feedback.csv", "w", newline="") as csvfile:
        csv_fieldnames = ["Intersections bâtiment projeté", "Autres"]
        csv_fieldnames.insert(0, "No OFS")
        csv_fieldnames.insert(1, "Commune")
        csv_fieldnames.append("Bâtiments manquants")
        writer = csv.DictWriter(csvfile, fieldnames=csv_fieldnames, delimiter=";")
        writer.writeheader()

        for gdenr in municipalities_registry.keys():
            print("Commune:", municipalities_registry[gdenr])
            if not int(gdenr) in unique_gdenr_list:
                # Municipalities not in error list, set everything to 0
                csvfile = writer.writerow(
                    {
                        "No OFS": gdenr,
                        "Commune": municipalities_registry[gdenr],
                        "Bâtiments manquants": "0",
                        "Intersections bâtiment projeté": "0",
                        "Autres": "0",
                    }
                )

            else:
                # Municipalities not in error list
                data = df[df["gdenr"] == int(gdenr)]

                print()
                print(f"data.shape[0] = {data.shape[0]} before")

                # Filters
                if "etat_mo" in filters:
                    # filter etat_mo "working in progress"
                    data = data[data["etat_mo"] != "En travail"]

                if "bat_proj" in filters:
                    # filter bat_roj
                    data = data[~((data["bat_proj"] == 1) | (data["gstat"] < 1004))]

                print(f"data.shape[0] = {data.shape[0]} after")

                municipality_feedback_filename = _excel_municipalityFeedback_filename_setter(gdenr, municipalities_registry[gdenr], dt, "feedback", "xlsx")
                municipalityFilePath = filepath.parent / municipality_feedback_filename

                # Error source
                for index, x in data.iterrows():
                    if x["out_issue_22"] == 1:
                        data.at[index, "Type erreur"] = "Bâtiment manquant"
                    elif x["bat_proj"] == 1:
                        data.at[index, "Type erreur"] = "Intersection bâtiment projeté"
                    else:
                        data.at[index, "Type erreur"] = "Autre"

                # Write statistics in csv
                csvfile = writer.writerow(
                    {
                        "No OFS": str(gdenr),
                        "Commune": municipalities_registry[gdenr],
                        "Bâtiments manquants": str(data[data["Type erreur"] == "Bâtiment manquant"].shape[0]),
                        "Intersections bâtiment projeté": str(data[data["Type erreur"] == "Intersection bâtiment projeté"].shape[0]),
                        "Autres": str(data[data["Type erreur"] == "Autre"].shape[0]),
                    }
                )

                # sorting data
                data.sort_values(by=["Type erreur", "Aide du SGRF", "egid"])

                # remove columns
                del data["out_issue_22"]
                del data["bat_proj"]
                del data["etat_mo"]

                if data.shape[0] > 0:
                    data.to_excel(municipalityFilePath, index=False)
                    remove_cols_by_names_openpyxl(municipalityFilePath, attributes_to_remove)

    return


def get_mail_template():
    template_relpath = os.environ["FEEDBACK_COMMUNES_v2_TEMPLATE_RELPATH"]
    file_loader = FileSystemLoader(template_relpath)
    env = Environment(loader=file_loader)
    return env.get_template("feedback_commune.html")


def get_email_registry():
    municipality_file_relpath = os.environ["FEEDBACK_COMMUNES_v2_MUNICIPALITY_FILE_RELPATH"]
    wb = load_workbook(municipality_file_relpath)
    ws = wb["Communes"]
    row = 2
    mail = {}
    while ws.cell(row, 1).value != None:
        mail[str(ws.cell(row, 2).value)] = ws.cell(row, 6).value
        row += 1
    return mail


def get_municipalities_registry(filepath, canton="NE"):
    df = pd.read_json(filepath)
    df = df[df["canton"] == canton]
    commune_registry = {}
    for x in df.itertuples(index=False):
        if x.id.isdigit():
            commune_registry[str(x.id)] = x.name
    return commune_registry


def send_mail(to, subject, content, files=[]):
    assert isinstance(to, list)
    assert isinstance(files, list)
    to = [value for value in to if value != None]
    to = [value for value in to if value != ""]

    if len(to) == 0:
        return

    msg = EmailMessage()
    msg["From"] = os.getenv("FEEDBACK_COMMUNES_v2_MAIL_EMAILADRESS_FROM")
    msg["To"] = ", ".join(to)
    msg["Date"] = formatdate(localtime=True)
    msg["Subject"] = subject
    msg["Body"] = ""
    msg.add_alternative(content, subtype="html")

    for f in files:
        # print(f)
        filename = f.name
        with open(f, "rb") as f:
            part = MIMEApplication(f.read(), Name=filename)
        # After the file is closed
        part["Content-Disposition"] = f'attachment; filename="{filename}"'
        msg.attach(part)

    s = smtplib.SMTP(os.getenv("FEEDBACK_COMMUNES_v2_MAIL_SMTP"))
    s.send_message(msg)
    s.quit()
    return


def send_mail_municipalities(feedback_path, municipality_filepath, testemail=None):

    feedback_csv_path = feedback_path.parent / f"{dt}-feedback.csv"

    municipalities_registry = get_municipalities_registry(municipality_filepath, canton="NE")

    jinja_tpl = get_mail_template()
    mails = get_email_registry()

    with open(feedback_csv_path, mode="r", newline="") as file:
        spamreader = csv.reader(file, delimiter=";")

        # parcourir les communes
        for row in spamreader:
            # ne pas tenir compte de la ligne des titres
            if row[0].isnumeric():
                # if int(row[0]) not in [6421, 6458]:
                #     continue

                total_error = sum([int(a) for a in row[2:]])

                data_tpl = {
                    "COMMUNE_NAME": row[1],
                    "NB_ERROR_INTERSECT_PROJECT_BUILDING": row[2],
                    "NB_ERROR_OTHER": row[3],
                    "NB_ERROR_ISSUE_22": row[4],
                    "ADMINISTRATOR_NAME": os.getenv("MAIL_COMMUNE_ADMINISTRATOR_NAME"),
                    "ADMINISTRATOR_EMAIL": os.getenv("MAIL_COMMUNE_ADMINISTRATOR_EMAIL"),
                }

                # remplir le template
                email_content = jinja_tpl.render(data_tpl)

                gdenr = row[0]

                municipality_feedback_filename = _excel_municipalityFeedback_filename_setter(gdenr, municipalities_registry[gdenr], dt, "feedback", "xlsx")
                attached_file = feedback_path.parent / municipality_feedback_filename

                # destinataire
                if testemail is not None:
                    to = re.split(";|,", testemail)
                else:
                    if int(gdenr) > 0:
                        to = [mails[row[0]]]
                        # continue
                    else:
                        # Canton de neuchâtel
                        to = [os.getenv("FEEDBACK_COMMUNES_v2_MAIL_ADMINISTRATOR_EMAIL")]

                # envoyer le mail avec la pièce jointe si existante
                attached_files = []
                if total_error > 0:
                    attached_files = [attached_file]
                # print(f"send_mail({to}, 'RegBL - apurement des données', email_content, files={attached_files})")
                send_mail(to, "RegBL - apurement des données", email_content, files=attached_files)  ################################

                mail_filename = _excel_municipalityFeedback_filename_setter(row[0], row[1], dt, "mail", "html")
                mail_path = filepath.parent / mail_filename
                with open(mail_path, "w", encoding="utf-8") as mail_file:
                    mail_file.write(email_content)

                # sdfs


def _auto_resize_worksheet(ws, max_col_width=50, min_col_width=10, base_row_height=15):
    """
    Ajuste automatiquement la largeur des colonnes et la hauteur des lignes
    selon le contenu de toutes les cellules de la feuille.

    :param ws: worksheet openpyxl
    :param max_col_width: largeur max d'une colonne (caractères Excel)
    :param min_col_width: largeur min d'une colonne
    :param base_row_height: hauteur d'une ligne (points)
    """

    # Stocke le max de caractères par colonne
    col_max_width = {}

    # Première passe : calcul des largeurs de colonnes
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue

            text = str(cell.value)
            lines = text.split("\n")
            longest_line = max(len(line) for line in lines)

            col_idx = cell.column
            col_max_width[col_idx] = max(col_max_width.get(col_idx, 0), longest_line)

            # Active le wrap text
            cell.alignment = Alignment(wrap_text=True)

    # Applique les largeurs de colonnes
    for col_idx, width in col_max_width.items():
        column_letter = ws.cell(row=1, column=col_idx).column_letter
        adjusted_width = min(max(width + 2, min_col_width), max_col_width)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Deuxième passe : calcul des hauteurs de lignes
    for row in ws.iter_rows():
        row_idx = row[0].row
        max_lines_in_row = 1

        for cell in row:
            if cell.value is None:
                continue

            text = str(cell.value)
            col_letter = cell.column_letter
            col_width = ws.column_dimensions[col_letter].width

            lines = text.split("\n")
            total_lines = 0
            for line in lines:
                wrapped = math.ceil(len(line) / col_width) if line else 1
                total_lines += wrapped

            max_lines_in_row = max(max_lines_in_row, total_lines)

        ws.row_dimensions[row_idx].height = max_lines_in_row * base_row_height


def launching_options(argv):
    opt = {
        "environ": None,
        "batproj_filter": None,
        "egidwhitelist_filter": None,
        "etatmo_filter": None,
        "sendmail": None,
        "testemail": None,
    }

    if len(sys.argv) == 1:
        # get arguments
        environ_input = input("Site internet ou intranet [INTER | INTRA (défaut)]: ")
        opt["environ"] = "INTER" if environ_input == "inter" else "INTRA"
        print(opt["environ"])

        batproj_filter_input = input("Filtrer les batiments projetés [oui (défaut) | non]: ")
        opt["batproj_filter"] = False if batproj_filter_input == "non" else True
        print("Oui" if opt["batproj_filter"] is True else "Non")

        egidwhitelist_filter_input = input("Filtrer les EGID avec la whiteliste [oui (défaut) | non]: ")
        opt["egidwhitelist_filter"] = False if egidwhitelist_filter_input == "non" else True
        print("Oui" if opt["egidwhitelist_filter"] is True else "Non")

        etatmo_filter_input = input("Filtrer les EGID selon l'état de la MO [oui (défaut) | non]: ")
        opt["etatmo_filter"] = False if etatmo_filter_input == "non" else True
        print("Oui" if opt["etatmo_filter"] is True else "Non")

        sendmail_input = input("Envoyer les e-mails [oui (défaut) | non]: ")
        opt["sendmail"] = False if sendmail_input == "non" else True
        print("Oui" if opt["sendmail"] is True else "Non")
        if opt["sendmail"] is True:
            opt["testemail"] = input("\tAdresse e-mail de test [envoi aux communes si vide, sinon à l'adresse e-mail spécifiée]: ")

    elif len(sys.argv) > 1:
        # autofill parameters
        if "--auto" in argv:
            opt["environ"] = "INTER"
            opt["batproj_filter"] = False
            opt["egidwhitelist_filter"] = True
            opt["etatmo_filter"] = True
            opt["sendmail"] = True
            opt["testemail"] = None
            print("internet ou intranet [INTER | INTRA (défaut)]: " + opt["environ"])
            print("Filtrer les batiments projetés [oui (défaut) | non]: " + ("Oui" if opt["batproj_filter"] is True else "Non"))
            print("Filtrer les EGID avec la whiteliste [oui (défaut) | non]: " + ("Oui" if opt["egidwhitelist_filter"] is True else "Non"))
            print("Filtrer les EGID selon l'état de la MO [oui (défaut) | non]: " + ("Oui" if opt["etatmo_filter"] is True else "Non"))
            print("Envoyer les e-mails [oui (défaut) | non]: " + ("Oui" if opt["sendmail"] is True else "Non"))

        else:
            raise ValueError("Invalid parameters ")

        # email to send to
        if "--prod" not in argv:
            opt["testemail"] = os.environ["FEEDBACK_COMMUNES_v2_EMAIL_MAILADRESS_TO_TEST"]
            print("\tAdresse e-mail de test [envoi aux communes si vide, sinon à l'adresse e-mail spécifiée]: " + opt["testemail"])
        else:
            print("\tAdresse e-mail de test [envoi aux communes si vide, sinon à l'adresse e-mail spécifiée]: ")

    print("===================================")
    print()
    print(f">> opt = {opt}")
    print()
    print("===================================")

    return opt


if __name__ == "__main__":
    opt = launching_options(sys.argv)

    # Fixed data
    url = os.environ["FEEDBACK_COMMUNES_v2_DOWNLOAD_SOURCE_LIST"]
    dt = datetime.strftime(datetime.now(), "%Y%m%d")
    zipPath = Path(os.environ["FEEDBACK_COMMUNES_v2_DATA_ROOT_PATH"]) / f"{dt}_monico.zip"
    json_dataFilename = os.environ["FEEDBACK_COMMUNES_v2_JSON_DATA_FILENAME"]
    excel_dataFilename = os.environ["FEEDBACK_COMMUNES_v2_EXCEL_DATA_FILENAME"]
    # downloaded municipalities file
    municipality_filename = os.environ["FEEDBACK_COMMUNES_v2_MUNICIPALITY_FILENAME"]

    # 1. Download file
    path = downloadFiles(url, zipPath, json_dataFilename)

    # 2. Get data from file
    filters = ["whitelist"] if opt["egidwhitelist_filter"] is True else []
    print(f"2. Filters = {filters}")
    filepath = getDataFrame(path, json_dataFilename, excel_dataFilename, filters=filters)

    # 3. Tune Excel file
    filepath = excelFineTuning(filepath)

    # 4. Remove extra columns and keep just what you need for FME
    list_of_columns_to_keep = os.getenv("FEEDBACK_COMMUNES_v2_COLUMNS_TO_KEEP").split(",")
    remove_cols_by_names_pandas(filepath, list_of_columns_to_keep)

    # 5. Call FME scripts (Bat proj / Etat MO / issue 22)
    callFME(filepath)

    # 6. save one file per administrative unit (municipalities)
    filters = []
    filters.append("bat_proj") if opt["batproj_filter"] is True else None
    filters.append("etat_mo") if opt["etatmo_filter"] is True else None
    print(f"6. Filters = {filters}")
    attributes_to_remove = ["gkode", "gkodn"]

    municipality_filepath = filepath.parent / municipality_filename
    exportMunicipalityFile(filepath, municipality_filepath, dt, filters=filters, attributes_to_remove=attributes_to_remove)

    # 7. Send mail to municipalities
    if opt["sendmail"] is True:
        send_mail_municipalities(filepath, municipality_filepath, testemail=opt["testemail"])
