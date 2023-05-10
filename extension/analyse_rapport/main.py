from pyproj import CRS, Transformer
import simplekml
import openpyxl
import re
import os

from dotenv import load_dotenv

load_dotenv(r'..\..\.env')

working_directory = os.environ['RAPPORT_ANALYZER_WORKING_DIRECTORY']
path_issue_solution = os.environ['RAPPORT_ANALYZER_ISSUE_SOLUTION_PATH']


crs_2056 = CRS.from_epsg(2056)
crs_4326 = CRS.from_epsg(4326)
TF = Transformer.from_crs(crs_2056, crs_4326, always_xy=True)


def get_issue_solution(filepath):
    if not os.path.exists(filepath):
        print("Le fichier {} n'existe pas".format(filepath))
        raise

    data = {}
    with open(filepath, 'r', encoding='utf-8') as f:
        for line in f:
            tmp = line.split(" ", maxsplit=1)
            data[tmp[0]] = tmp[1][:-1]
    return data


def _get_issue(error, issue_solution):
    c_split = error.split("</br>")
    solutions = []
    issues_id = []
    for c_part in c_split:
        tmp = c_part[0:2]
        issues_id.append(tmp)
        if tmp.isnumeric():
            if tmp in issue_solution.keys():
                solutions.append(issue_solution[tmp])
            else:
                solutions.append("Non défini")
    issues_id_concat = "-".join(issues_id)
    if issues_id_concat in issue_solution:
        issue = issue_solution[issues_id_concat]
    else:
        issue =  "; ".join(solutions)
    return issue


def _liste_5(ws, row, issue_solution):
    coord = [ws.cell(row, 5).value, ws.cell(row, 6).value]
    # coord = ws.cell(row, 5).value.split(' ')
    color = simplekml.Color.green
    nom = str(ws.cell(row, 1).value) + ' ' + _get_issue(ws.cell(row, 7).value, issue_solution)
    return (coord, color, nom)


def _liste_6(ws, row, issue_solution):
    coord = [ws.cell(row, 5).value, ws.cell(row, 6).value]
    # coord = ws.cell(row, 5).value.split(' ')
    color = simplekml.Color.blue
    nom = str(ws.cell(row, 1).value) + ' ' + _get_issue(ws.cell(row, 7).value, issue_solution)
    return (coord, color, nom)


def _liste_2remove(ws, row, issue_solution):
    coord = [ws.cell(row, 5).value, ws.cell(row, 6).value]
    # coord = ws.cell(row, 5).value.split(' ')
    color = simplekml.Color.red
    nom = str(ws.cell(row, 1).value) + ' ' + _get_issue(ws.cell(row, 7).value, issue_solution)
    return (coord, color, nom)


def _liste_bat_manquants(ws, row, issue_solution):
    coord = [ws.cell(row, 1).value, ws.cell(row, 2).value]
    color = simplekml.Color.yellow
    nom = 'bât. manquant ({}, {})'.format(ws.cell(row, 1).value, ws.cell(row, 2).value)
    return (coord, color, nom)


def _liste_coordonnees_fausses_manquants(ws, row, issue_solution):
    coord = [ws.cell(row, 5).value, ws.cell(row, 6).value]
    color = simplekml.Color.black
    nom = 'coord fausses ({})'.format(ws.cell(row, 7).value)
    return (coord, color, nom)


def xlsx2kml(filepath, sheetname, kml_filename):
    kml = simplekml.Kml()
    
    issue_solution = get_issue_solution(path_issue_solution)


    wb=openpyxl.load_workbook(filepath, data_only=True)
    ws = wb[sheetname]
    
    c = 0
    i = 1
    actual_list_fct = None
    go_further = False
    while i<10000:
        i += 1
        
        # Si la cellule est nulle
        if ws.cell(i, 1).value is None:
            c += 1
        
            # Fin du document
            if c > 5:
                break
        
            continue

        if ws.cell(i, 2).value is not None and str(ws.cell(i, 2).value).isnumeric() and int(ws.cell(i, 2).value) < 1060:
            continue

        # La cellule n'est pas nulle
        c = 0

        print(i, ws.cell(i, 1).value)
        # Initialiser la fonction actuelle
        if str(ws.cell(i, 1).value).startswith('Liste 5 '):
            actual_list_fct = _liste_5
            go_further = True
            i += 1    # éviter la ligne de titres
            continue
        elif str(ws.cell(i, 1).value).startswith('Liste 6 '):
            actual_list_fct = _liste_6
            go_further = True
            i += 1    # éviter la ligne de titres
            continue
        elif str(ws.cell(i, 1).value).startswith('Gebäude in der AV zu löschen / Bâtiment à supprimer dans la MO'):
            actual_list_fct = _liste_2remove
            go_further = True
            i += 1    # éviter la ligne de titres
            continue
        elif str(ws.cell(i, 1).value).startswith('Fehlende Gebäude / Bâtiments manquants (issue 22)'):
            actual_list_fct = _liste_bat_manquants
            go_further = True
            i += 1    # éviter la ligne de titres
            continue
        elif str(ws.cell(i, 1).value).startswith('Falsche Koordinaten / Coordonnées à corriger'):
            actual_list_fct = _liste_coordonnees_fausses_manquants
            go_further = True
            i += 1    # éviter la ligne de titres
            continue

        if go_further is True:
            coord, color, nom = actual_list_fct(ws, i, issue_solution)
            
            if not 'Ne rien faire' in nom:
                pnt = kml.newpoint(
                    name = nom,
                    coords = [TF.transform(float(coord[0]), float(coord[1]))],
                )
                pnt.style.labelstyle.color = color

        kml.save(kml_filename)
    return




if __name__ == "__main__":
    files = os.listdir(working_directory)
    for filename in files:
        if re.match(r'6[4,5]{1}[\d]{2}.*\.xlsx', filename) is not None:
            no_commune = filename.split('.')[0].split('_')[0]
            print(no_commune)
            input_filepath = os.path.join(working_directory, filename)
            kml_filename = no_commune + ".kml"
            output_filepath = os.path.join(working_directory, kml_filename)
            xlsx2kml(input_filepath, no_commune, output_filepath)
