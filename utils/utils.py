from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from datetime import datetime, date
import requests
import os
import re
import shutil
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker


def _findRowIndex(ws, searchTerm, column_id=1, limit=1e3, egidextfilter=False):
    index = None

    row_id = 1
    while row_id < limit:
        if ws.cell(row_id, column_id).value == searchTerm:
            if egidextfilter is True and ws.cell(row_id, 4).value >= 500000000:
                index = row_id
                break
            else:
                index = row_id
                break
        row_id += 1

    return index


def _findColumnIndex(ws, searchTerm, row_id=1, limit=1e3):
    index = None

    column_id = 1
    while column_id < limit:
        if ws.cell(row_id, column_id).value == searchTerm:
            index = column_id
            break
        column_id += 1

    return index


def _removeRows(ws, searchTerm, row_id=1, column_id=1, limit=1e3):
    while row_id < limit and ws.cell(row_id, column_id).value is not None:
        if ws.cell(row_id, column_id).value == searchTerm:
            row_id += 1
        else:
            ws.delete_rows(row_id)
    return


# def _getIssue22OfCommune(wb, commune_id, issue22_list):
#     ws = wb.create_sheet('ISSUE22')
#     currentLine_idx = 1

#     if issue22_list == []:
#         return

#     # title line
#     for j, title in enumerate(issue22_list[0].keys()):
#         ws.cell(currentLine_idx,j+1).value = title

#     currentLine_idx += 1

#     for i22 in issue22_list:
#         if i22['COM_FOSNR'] == commune_id:
#             ws.cell(currentLine_idx,1).value = i22['COM_FOSNR']
#             ws.cell(currentLine_idx,2).value = i22['AV_SOURCE']
#             ws.cell(currentLine_idx,3).value = i22['AV_TYPE']
#             ws.cell(currentLine_idx,4).value = i22['ISSUE']
#             ws.cell(currentLine_idx,5).value = i22['ISSUE_CATEGORY']
#             ws.cell(currentLine_idx,6).value = i22['BDG_E']
#             ws.cell(currentLine_idx,7).value = i22['BDG_N']
#             ws.cell(currentLine_idx,8).hyperlink = os.environ['FEEDBACK_COMMUNES_URL_CONSULTATION_' + environ + '_ISSUE_22_SITN_COORD'].format(i22['BDG_E'],i22['BDG_N'])
#             ws.cell(currentLine_idx,8).value = 'sitn.ne.ch'
#             ws.cell(currentLine_idx,8).style = 'Hyperlink'
#             currentLine_idx += 1

#     return wb


def _get_issue(error, issue_solution):
    c_split = error.split("</br>")
    solutions = []
    issues_id = []
    for c_part in c_split:
        tmp = c_part[0:2]
        issues_id.append(tmp)
        if tmp.isnumeric():
            if tmp in issue_solution.keys():
                solutions.append(issue_solution[tmp])
            else:
                solutions.append("Non défini")
    issues_id_concat = "-".join(issues_id)
    if issues_id_concat in issue_solution:
        issue = issue_solution[issues_id_concat]
    else:
        issue = "; ".join(solutions)
    return issue


def get_issue_solution(filepath):
    if not os.path.exists(filepath):
        print("Le fichier {} n'existe pas".format(filepath))
        raise

    data = {}
    with open(filepath, "r", encoding="utf-8") as f:
        for line in f:
            tmp = line.split(" ", maxsplit=1)
            data[tmp[0]] = tmp[1][:-1]
    return data


def loadCommunesOFS(filepath):
    communes_ofs = {}

    wb_source = load_workbook(filepath)
    ws_source = wb_source["Communes"]

    i = _findRowIndex(ws_source, "NE", column_id=2) - 1

    while i < 1e2:
        i += 1

        if ws_source.cell(i, 2).value is None:
            break

        if ws_source.cell(i, 2).value != "NE":
            continue

        communes_ofs[ws_source.cell(i, 3).value] = ws_source.cell(i, 4).value

    return communes_ofs


def downloadListeCantonNeuchatel(path=None, batprojtreat=False):
    if path is None or not os.path.exists(path):
        path = os.environ["FEEDBACK_COMMUNES_WORKING_DIR"]

    r = requests.get(os.environ["STATIC_URL_FEEBDACK_CANTON"], allow_redirects=True)

    filename = datetime.strftime(datetime.now(), "%Y%m%d_Listes_NE.xlsx")

    filepath = os.path.join(path, filename)
    open(filepath, "wb").write(r.content)
    if batprojtreat is True:
        tempfilepath = os.path.join(path, "temp_" + filename)
        open(tempfilepath, "wb").write(r.content)

        fme_command = 'fme "{}" --SourceDataset_POSTGIS "{}" --SourceDataset_XLSXR "{}" --DestDataset_XLSXW "{}"'.format(os.environ["FEEDBACK_COMMUNES_LISTES_BATPROJ_FILTER_FME_PATH"], os.environ["FEEDBACK_COMMUNES_LISTES_BATPROJ_FILTER_FME_USER"], tempfilepath, filepath)
        print('\ncommand: {}"\n'.format(fme_command))
        os.system(fme_command)

        os.remove(tempfilepath)

    return filepath


def downloadIssue22CantonNeuchatel(path=None):
    if path is None or not os.path.exists(path):
        path = os.environ["FEEDBACK_COMMUNES_WORKING_DIR"]

    r = requests.get(os.environ["FEEDBACK_COMMUNES_URL_ISSUE_22_CANTON"], allow_redirects=True)

    filename = datetime.strftime(datetime.now(), "%Y%m%d_issue22_NE.xlsx")

    filepath = os.path.join(path, filename)
    open(filepath, "wb").write(r.content)
    tempfilepath = os.path.join(path, "temp_" + filename)
    open(tempfilepath, "wb").write(r.content)

    # Filter projected buildings
    fme_command = 'fme "{}" --SourceDataset_POSTGIS "{}" --SourceDataset_XLSXR "{}" --DestDataset_XLSXW "{}'.format(os.environ["FEEDBACK_COMMUNES_ISSUE_22_FME_PATH"], os.environ["FEEDBACK_COMMUNES_ISSUE_22_FME_USER"], tempfilepath, filepath)
    print('\ncommand: {}"\n'.format(fme_command))
    os.system(fme_command)

    os.remove(tempfilepath)

    wb = load_workbook(filepath)
    ws = wb["NE"]

    issue22_list = []
    i = 1
    while i < 1e5 and ws.cell(i, 1) is not None:
        if ws.cell(i, 5).value == 22:
            issue22_list.append(
                {"COM_FOSNR": ws.cell(i, 1).value, "AV_SOURCE": ws.cell(i, 2).value, "AV_TYPE": ws.cell(i, 3).value, "ISSUE": ws.cell(i, 4).value, "ISSUE_CATEGORY": ws.cell(i, 5).value, "BDG_E": ws.cell(i, 6).value, "BDG_N": ws.cell(i, 7).value, "DESIGNATION_MO": ws.cell(i, 8).value}
            )

        i += 1

    return (issue22_list, filepath)


def cleanWorkingDirectory(path=None):
    if path is not None and os.path.exists(path):
        shutil.rmtree(path)
    return


def _createExcelTable(ws, table_name, ref):
    tab = Table(displayName=table_name, ref=ref)

    style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)

    tab.tableStyleInfo = style

    ws.add_table(tab)

    return


def _tableTitleGenerator(ws, sectionTitle, tableHeader, line_idx):
    line_idx += 1

    ws.cell(line_idx, 1).value = sectionTitle
    ws.cell(line_idx, 1).style = "Headline 1"

    line_idx += 1
    table_start_row = line_idx

    for i, header in enumerate(tableHeader):
        ws.cell(line_idx, i + 1).value = header

    line_idx += 1

    return line_idx, table_start_row


def _getDateFromExcel(excel_date):
    dt = None
    if excel_date is not None:
        dt = datetime.fromordinal(datetime(1900, 1, 1).toordinal() + excel_date - 2)
        tt = dt.timetuple()
        print(dt)
        print(tt)
    return dt


def getEGIDWhitelistSGRF(path=None):
    if path is None:
        os.environ["FEEDBACK_COMMUNES_EGID_WHITELIST_CONTROLS_PATH"]

    whitelist_sgrf = {}
    if os.path.exists(path):
        wb_whitelist = load_workbook(path)
        ws_whitelist = wb_whitelist.active
        i = 2
        while ws_whitelist.cell(i, 1).value is not None:  # il y a un EGID
            if ws_whitelist.cell(i, 4).value is None:  # il n'y a pas de date de sortie
                whitelist_sgrf[ws_whitelist.cell(i, 1).value] = {
                    "egid": ws_whitelist.cell(i, 1).value,
                    "commune": ws_whitelist.cell(i, 2).value,
                    "date_entree": date.strftime(ws_whitelist.cell(i, 3).value, "%d.%m.%Y") if ws_whitelist.cell(i, 3).value is not None else "",
                    "date_sortie": date.strftime(ws_whitelist.cell(i, 4).value, "%d.%m.%Y") if ws_whitelist.cell(i, 4).value is not None else "",
                    "remarque": ws_whitelist.cell(i, 5).value or "",
                    "xlsx_row_id": i,
                }
                i += 1
    return whitelist_sgrf


def _log_whitelisted_egid(egid, egidwhitelistsgrf, ws_whitelist, today, commune_name):
    filepath = os.path.join(os.environ["FEEDBACK_COMMUNES_WORKING_DIR"], today, f"{today}_whitelisted.csv")
    with open(filepath, "a+") as file:
        if os.stat(filepath).st_size == 0:
            file.write(f"EGID;Commune;Date entree;Remarque\n")

        file.write(f'{egid};{commune_name};{egidwhitelistsgrf["date_entree"]};{egidwhitelistsgrf["remarque"]}\n')

    # Ecrire la date dans le fichier whitelist
    if ws_whitelist is not None:
        ws_whitelist.cell(egidwhitelistsgrf["xlsx_row_id"], 6).value = f"{today[6:]:2}.{today[4:6]:2}.{today[:4]:4}"

    return


def generateCommuneErrorFile(commune_id, commune_name, feedback_canton_filepath, issue22_list, issue_solution, today=datetime.strftime(datetime.now(), "%Y%m%d"), environ="INTER", egidextfilter=False, log=False, egidwhitelist=[], whitelist_path=None):
    # get active sheet of whitelist
    ws_whitelist = None
    if whitelist_path is not None and os.path.exists(whitelist_path):
        wb_whitelist = load_workbook(whitelist_path)
        ws_whitelist = wb_whitelist.active

    # copy canton_file to commune_file
    feedback_commune_filename = "_".join([str(commune_id), commune_name.replace(" ", "_"), "feedback", today]) + ".xlsx"
    feedback_commune_filepath = os.path.join(os.environ["FEEDBACK_COMMUNES_WORKING_DIR"], today, feedback_commune_filename)
    shutil.copy2(feedback_canton_filepath, feedback_commune_filepath)

    nb_errors_by_list = {}
    nb_errors_by_list["Commune_id"] = commune_id
    nb_errors_by_list["Commune"] = commune_name

    feedback_commune = {"commune_id": commune_id, "commune_nom": commune_name}

    wb = load_workbook(feedback_commune_filepath)
    if log:
        print(f"${feedback_commune_filepath} loaded")

    # remove canton sheet
    wb.remove(wb["Cantons"])

    # remove other communes in communes sheet
    ws = wb["Communes"]
    row_i = _findRowIndex(ws, "Commune", column_id=4)
    _removeRows(ws, commune_id, row_id=row_i + 2, column_id=3, limit=1e2)

    # Create resume sheet
    ws2 = wb.create_sheet("resume")

    #####################
    #  INFOS GENERALES
    #####################
    if log:
        print("Infos gen - start")
    ws = wb["Communes"]
    line_i = _findRowIndex(ws, commune_id, 3)

    ws2.cell(1, 1).value = commune_id
    ws2.cell(1, 1).style = "Title"
    ws2.cell(1, 2).value = commune_name
    ws2.cell(1, 2).style = "Title"
    ws2.cell(1, 4).value = datetime.strftime(datetime.now(), "%d.%m.%Y")
    ws2.cell(1, 6).value = "Explicatif sur la manière de traiter les incohérences"
    ws2.cell(1, 6).hyperlink = "https://www.housing-stat.ch/files/Traitement_erreurs_FR.pdf"
    ws2.cell(1, 6).style = "Hyperlink"
    ws2.cell(3, 1).value = "Nombre de bâtiments: "
    ws2.cell(3, 5).value = ws.cell(line_i, 5).value
    ws2.cell(4, 1).value = "Bâtiments manquants: "
    ws2.cell(4, 5).value = ws.cell(line_i, 9).value
    ws2.cell(5, 1).value = "Total erreurs 1-6: "
    ws2.cell(5, 5).value = sum([ws.cell(line_i, 12 + k * 5).value for k in range(6)])
    ws2.cell(5, 6).value = "soit"
    ws2.cell(5, 7).value = "{0:.2f}%".format(ws.cell(line_i, 41).value * 100)
    ws2.cell(5, 8).value = "des bâtiments saisis dans le RegBL."
    ws2.cell(6, 1).value = "Fichier KML"
    ws2.cell(6, 5).value = ws.cell(line_i, 8).value
    ws2.cell(6, 5).hyperlink = ws.cell(line_i, 8).hyperlink
    ws2.cell(6, 5).style = "Hyperlink"
    ws2.cell(6, 6).value = "https://data.geo.admin.ch/ch.bfs.gebaeude_wohnungs_register/address/NE/{}_bdg_erw.kml".format(commune_id)
    ws2.cell(6, 6).hyperlink = "https://data.geo.admin.ch/ch.bfs.gebaeude_wohnungs_register/address/NE/{}_bdg_erw.kml".format(commune_id)
    ws2.cell(6, 6).style = "Hyperlink"

    ws2.cell(8, 1).value = "Bâtiments sans usage d'habitation (déjà dans le RegBL)"
    ws2.cell(8, 1).style = "Headline 1"
    ws2.cell(9, 1).value = "Nombre"
    [ws2.merge_cells(start_row=9, start_column=2 + k * 2, end_row=9, end_column=3 + k * 2) for k in range(3)]
    ws2.cell(9, 2).value = "avec GKLAS"
    ws2.cell(9, 4).value = "avec GBAUP"
    ws2.cell(9, 6).value = "avec GKLAS + GBAUP"
    ws2.cell(9, 8).value = "Update: {}".format(ws.cell(1, 2).value.replace("Etat: ", ""))
    ws2.cell(10, 1).value = ws.cell(line_i, 42).value
    ws2.cell(10, 2).value = ws.cell(line_i, 43).value
    ws2.cell(10, 3).value = "{0:.0f}%".format(ws.cell(line_i, 44).value * 100)
    ws2.cell(10, 4).value = ws.cell(line_i, 45).value
    ws2.cell(10, 5).value = "{0:.0f}%".format(ws.cell(line_i, 46).value * 100)
    ws2.cell(10, 6).value = ws.cell(line_i, 47).value
    ws2.cell(10, 7).value = "{0:.0f}%".format(ws.cell(line_i, 48).value * 100)
    ws2.cell(10, 8).value = "(GKAT 1060, tous)"
    ws2.cell(11, 1).value = ws.cell(line_i, 49).value
    ws2.cell(11, 2).value = ws.cell(line_i, 50).value
    ws2.cell(11, 3).value = "{0:.0f}%".format(ws.cell(line_i, 51).value * 100)
    ws2.cell(11, 4).value = ws.cell(line_i, 52).value
    ws2.cell(11, 5).value = "{0:.0f}%".format(ws.cell(line_i, 53).value * 100)
    ws2.cell(11, 6).value = ws.cell(line_i, 54).value
    ws2.cell(11, 7).value = "{0:.0f}%".format(ws.cell(line_i, 55).value * 100)
    ws2.cell(11, 8).value = "(GKAT 1060, GAREA > 30m2)"

    ws2_line_i = 12
    if log:
        print("Infos gen - end")

    (table_start_row, table_end_row) = (0, 0)

    #####################
    #  LISTE 1
    #####################
    if log:
        print("Liste 1 - start")
    ws = wb["Liste 1"]
    ws_line_i = _findRowIndex(ws, commune_id, column_id=2, limit=1e4, egidextfilter=egidextfilter)

    table_start_row = None

    if ws_line_i is not None:
        while ws.cell(ws_line_i, 2).value is not None:
            if ws.cell(ws_line_i, 2).value == commune_id:
                if table_start_row is None:
                    tableHeader = ["EGID", "STRNAME", "DEINR", "PLZ4", "PLZNAME"]
                    ws2_line_i, table_start_row = _tableTitleGenerator(ws2, "LISTE 1 - Bâtiments sans coordonnées", tableHeader, ws2_line_i)

                if ws.cell(ws_line_i, 4).value in egidwhitelist.keys():
                    _log_whitelisted_egid(ws.cell(ws_line_i, 4).value, egidwhitelist[ws.cell(ws_line_i, 4).value], ws_whitelist, today, commune_name)

                elif egidextfilter is False or (egidextfilter is True and ws.cell(ws_line_i, 4).value < 500000000):
                    ws2.cell(ws2_line_i, 1).value = ws.cell(ws_line_i, 4).value
                    ws2.cell(ws2_line_i, 2).value = ws.cell(ws_line_i, 11).value
                    ws2.cell(ws2_line_i, 3).value = ws.cell(ws_line_i, 12).value
                    ws2.cell(ws2_line_i, 4).value = ws.cell(ws_line_i, 13).value
                    ws2.cell(ws2_line_i, 5).value = ws.cell(ws_line_i, 15).value
                    ws2_line_i += 1

            ws_line_i += 1

        table_end_row = ws2_line_i - 1

    if table_start_row is not None and table_end_row > table_start_row:
        ref = "A" + str(table_start_row) + ":" + get_column_letter(5) + str(table_end_row)
        _createExcelTable(ws2, "Liste1", ref)

    if table_start_row is None:
        nb_errors_by_list["Liste_1"] = 0
    else:
        nb_errors_by_list["Liste_1"] = table_end_row - table_start_row
    if log:
        print("Liste 1 - end")

    #####################
    #  LISTE 2
    #####################
    if log:
        print("Liste 2 - start")
    ws = wb["Liste 2"]
    ws_line_i = _findRowIndex(ws, commune_id, column_id=2, limit=1e4, egidextfilter=egidextfilter)

    table_start_row = None

    if ws_line_i is not None:
        while ws.cell(ws_line_i, 2).value is not None:
            statut_base_mo = ws.cell(ws_line_i, 19).value if ws.cell(ws_line_i, 19).value is not None else ""
            if ws.cell(ws_line_i, 2).value == commune_id and not re.search("En travail", statut_base_mo, re.IGNORECASE):
                if table_start_row is None:
                    tableHeader = ["EGID", "Adresse", "GKODE", "GKODN"]
                    ws2_line_i, table_start_row = _tableTitleGenerator(ws2, "LISTE 2 - Coordonnées en dehors de la commune", tableHeader, ws2_line_i)

                if ws.cell(ws_line_i, 4).value in egidwhitelist.keys():
                    _log_whitelisted_egid(ws.cell(ws_line_i, 4).value, egidwhitelist[ws.cell(ws_line_i, 4).value], ws_whitelist, today, commune_name)

                elif egidextfilter is False or (egidextfilter is True and ws.cell(ws_line_i, 4).value < 500000000):
                    ws2.cell(ws2_line_i, 1).value = ws.cell(ws_line_i, 4).value
                    ws2.cell(ws2_line_i, 1).hyperlink = os.environ["FEEDBACK_COMMUNES_URL_CONSULTATION_" + environ + "_ISSUE_22_SITN_COORD"].format(ws.cell(ws_line_i, 11).value, ws.cell(ws_line_i, 12).value)
                    ws2.cell(ws2_line_i, 1).style = "Hyperlink"
                    ws2.cell(ws2_line_i, 2).value = ws.cell(ws_line_i, 5).value
                    ws2.cell(ws2_line_i, 3).value = ws.cell(ws_line_i, 11).value
                    ws2.cell(ws2_line_i, 4).value = ws.cell(ws_line_i, 12).value
                    ws2_line_i += 1

            ws_line_i += 1

        table_end_row = ws2_line_i - 1

    if table_start_row is not None and table_end_row > table_start_row:
        ref = "A" + str(table_start_row) + ":" + get_column_letter(4) + str(table_end_row)
        _createExcelTable(ws2, "Liste2", ref)

    if table_start_row is None:
        nb_errors_by_list["Liste_2"] = 0
    else:
        nb_errors_by_list["Liste_2"] = table_end_row - table_start_row
    if log:
        print("Liste 2 - end")

    #####################
    #  LISTE 3
    #####################
    if log:
        print("Liste 3 - start")
    ws = wb["Liste 3"]
    ws_line_i = _findRowIndex(ws, commune_id, column_id=2, limit=1e4, egidextfilter=egidextfilter)

    table_start_row = None

    if ws_line_i is not None:
        while ws.cell(ws_line_i, 2).value is not None:
            statut_base_mo = ws.cell(ws_line_i, 27).value if ws.cell(ws_line_i, 27).value is not None else ""
            if ws.cell(ws_line_i, 2).value == commune_id and not re.search("En travail", statut_base_mo, re.IGNORECASE):
                if table_start_row is None:
                    tableHeader = ["EGID", "PLZ4 RegBL", "PLZ4_Name RegBL", "PLZ4 MO", "PLZ4_Name MO"]
                    ws2_line_i, table_start_row = _tableTitleGenerator(ws2, "LISTE 3 - Divergence de NPA", tableHeader, ws2_line_i)

                if ws.cell(ws_line_i, 4).value in egidwhitelist.keys():
                    _log_whitelisted_egid(ws.cell(ws_line_i, 4).value, egidwhitelist[ws.cell(ws_line_i, 4).value], ws_whitelist, today, commune_name)

                elif egidextfilter is False or (egidextfilter is True and ws.cell(ws_line_i, 4).value < 500000000):
                    ws2.cell(ws2_line_i, 1).value = ws.cell(ws_line_i, 4).value
                    ws2.cell(ws2_line_i, 1).hyperlink = os.environ["FEEDBACK_COMMUNES_URL_CONSULTATION_" + environ + "_ISSUE_22_SITN_COORD"].format(ws.cell(ws_line_i, 20).value, ws.cell(ws_line_i, 21).value)
                    ws2.cell(ws2_line_i, 1).style = "Hyperlink"
                    ws2.cell(ws2_line_i, 2).value = ws.cell(ws_line_i, 8).value
                    ws2.cell(ws2_line_i, 3).value = ws.cell(ws_line_i, 9).value
                    ws2.cell(ws2_line_i, 4).value = ws.cell(ws_line_i, 11).value
                    ws2.cell(ws2_line_i, 5).value = ws.cell(ws_line_i, 12).value
                    ws2_line_i += 1

            ws_line_i += 1

        table_end_row = ws2_line_i - 1

    if table_start_row is not None and table_end_row > table_start_row:
        ref = "A" + str(table_start_row) + ":" + get_column_letter(5) + str(table_end_row)
        _createExcelTable(ws2, "Liste3", ref)

    if table_start_row is None:
        nb_errors_by_list["Liste_3"] = 0
    else:
        nb_errors_by_list["Liste_3"] = table_end_row - table_start_row
    if log:
        print("Liste 3 - end")

    #####################
    #  LISTE 4
    #####################
    if log:
        print("Liste 4 - start")
    ws = wb["Liste 4"]
    ws_line_i = _findRowIndex(ws, commune_id, column_id=2, limit=1e4, egidextfilter=egidextfilter)

    table_start_row = None

    if ws_line_i is not None:
        while ws.cell(ws_line_i, 2).value is not None:
            statut_base_mo = ws.cell(ws_line_i, 26).value if ws.cell(ws_line_i, 26).value is not None else ""
            if ws.cell(ws_line_i, 2).value == commune_id and not re.search("En travail", statut_base_mo, re.IGNORECASE):
                if table_start_row is None:
                    tableHeader = ["EGID", "GKAT", "GPARZ", "GEBNR", "STRNAME", "DEINR", "PLZ4", "GBEZ", "BUR / REE"]
                    ws2_line_i, table_start_row = _tableTitleGenerator(ws2, "LISTE 4 - Doublets d'adresses", tableHeader, ws2_line_i)

                if ws.cell(ws_line_i, 4).value in egidwhitelist.keys():
                    _log_whitelisted_egid(ws.cell(ws_line_i, 4).value, egidwhitelist[ws.cell(ws_line_i, 4).value], ws_whitelist, today, commune_name)

                elif egidextfilter is False or (egidextfilter is True and ws.cell(ws_line_i, 4).value < 500000000):
                    ws2.cell(ws2_line_i, 1).value = ws.cell(ws_line_i, 4).value
                    ws2.cell(ws2_line_i, 1).hyperlink = os.environ["FEEDBACK_COMMUNES_URL_CONSULTATION_" + environ + "_ISSUE_22_SITN_COORD"].format(ws.cell(ws_line_i, 7).value, ws.cell(ws_line_i, 8).value)
                    ws2.cell(ws2_line_i, 1).style = "Hyperlink"
                    ws2.cell(ws2_line_i, 2).value = ws.cell(ws_line_i, 6).value
                    ws2.cell(ws2_line_i, 3).value = ws.cell(ws_line_i, 22).value
                    ws2.cell(ws2_line_i, 4).value = ws.cell(ws_line_i, 23).value
                    ws2.cell(ws2_line_i, 5).value = ws.cell(ws_line_i, 11).value
                    ws2.cell(ws2_line_i, 6).value = ws.cell(ws_line_i, 12).value
                    ws2.cell(ws2_line_i, 7).value = ws.cell(ws_line_i, 13).value
                    ws2.cell(ws2_line_i, 8).value = ws.cell(ws_line_i, 14).value
                    ws2.cell(ws2_line_i, 9).value = ws.cell(ws_line_i, 24).value
                    ws2_line_i += 1

            ws_line_i += 1

        table_end_row = ws2_line_i - 1

    if table_start_row is not None and table_end_row > table_start_row:
        ref = "A" + str(table_start_row) + ":" + get_column_letter(9) + str(table_end_row)
        _createExcelTable(ws2, "Liste4", ref)

    if table_start_row is None:
        nb_errors_by_list["Liste_4"] = 0
    else:
        nb_errors_by_list["Liste_4"] = table_end_row - table_start_row
    if log:
        print("Liste 4 - end")

    #####################
    #  LISTE 5
    #####################
    if log:
        print("Liste 5 - start")
    ws = wb["Liste 5"]
    ws_line_i = _findRowIndex(ws, commune_id, column_id=2, limit=1e4, egidextfilter=egidextfilter)

    table_start_row = None

    if ws_line_i is not None:
        while ws.cell(ws_line_i, 2).value is not None:
            statut_base_mo = ws.cell(ws_line_i, 14).value if ws.cell(ws_line_i, 14).value is not None else ""
            if ws.cell(ws_line_i, 2).value == commune_id and ws.cell(ws_line_i, 13).value != "bat_proj" and not re.search("En travail", statut_base_mo, re.IGNORECASE):
                if table_start_row is None:
                    tableHeader = ["EGID", "GKAT", "GKLAS", "GSTAT", "GKODE", "GKODN", "ISSUE", "RESOLUTION_SGRF"]
                    ws2_line_i, table_start_row = _tableTitleGenerator(ws2, "LISTE 5 - Définition du bâtiment", tableHeader, ws2_line_i)

                if ws.cell(ws_line_i, 4).value in egidwhitelist.keys():
                    _log_whitelisted_egid(ws.cell(ws_line_i, 4).value, egidwhitelist[ws.cell(ws_line_i, 4).value], ws_whitelist, today, commune_name)

                elif egidextfilter is False or (egidextfilter is True and ws.cell(ws_line_i, 4).value < 500000000):
                    ws2.cell(ws2_line_i, 1).value = ws.cell(ws_line_i, 4).value
                    (coord_e, coord_n) = ws.cell(ws_line_i, 9).value.split(" ")
                    ws2.cell(ws2_line_i, 1).hyperlink = os.environ["FEEDBACK_COMMUNES_URL_CONSULTATION_" + environ + "_ISSUE_22_SITN_COORD"].format(coord_e, coord_n)
                    ws2.cell(ws2_line_i, 1).style = "Hyperlink"
                    ws2.cell(ws2_line_i, 2).value = ws.cell(ws_line_i, 5).value
                    ws2.cell(ws2_line_i, 3).value = ws.cell(ws_line_i, 6).value
                    ws2.cell(ws2_line_i, 4).value = ws.cell(ws_line_i, 7).value
                    ws2.cell(ws2_line_i, 5).value = coord_e
                    ws2.cell(ws2_line_i, 6).value = coord_n
                    ws2.cell(ws2_line_i, 7).value = ws.cell(ws_line_i, 12).value
                    ws2.cell(ws2_line_i, 8).value = _get_issue(ws.cell(ws_line_i, 12).value, issue_solution)
                    ws2_line_i += 1

            ws_line_i += 1

        table_end_row = ws2_line_i - 1

    if table_start_row is not None and table_end_row > table_start_row:
        ref = "A" + str(table_start_row) + ":" + get_column_letter(8) + str(table_end_row)
        _createExcelTable(ws2, "Liste5", ref)

    if table_start_row is None:
        nb_errors_by_list["Liste_5"] = 0
    else:
        nb_errors_by_list["Liste_5"] = table_end_row - table_start_row
    if log:
        print("Liste 5 - end")

    #####################
    #  LISTE 6
    #####################
    if log:
        print("Liste 6 - start")
    ws = wb["Liste 6"]
    ws_line_i = _findRowIndex(ws, commune_id, column_id=2, limit=1e4, egidextfilter=egidextfilter)

    table_start_row = None

    if ws_line_i is not None:
        while ws.cell(ws_line_i, 2).value is not None:
            statut_base_mo = ws.cell(ws_line_i, 14).value if ws.cell(ws_line_i, 14).value is not None else ""
            if ws.cell(ws_line_i, 2).value == commune_id and ws.cell(ws_line_i, 13).value != "bat_proj" and not re.search("En travail", statut_base_mo, re.IGNORECASE):
                if table_start_row is None:
                    tableHeader = ["EGID", "GKAT", "GKLAS", "GSTAT", "GKODE", "GKODN", "ISSUE", "RESOLUTION_SGRF"]
                    ws2_line_i, table_start_row = _tableTitleGenerator(ws2, "Liste 6 - Catégorie du bâtiment", tableHeader, ws2_line_i)

                if ws.cell(ws_line_i, 4).value in egidwhitelist.keys():
                    _log_whitelisted_egid(ws.cell(ws_line_i, 4).value, egidwhitelist[ws.cell(ws_line_i, 4).value], ws_whitelist, today, commune_name)

                elif egidextfilter is False or (egidextfilter is True and ws.cell(ws_line_i, 4).value < 500000000):
                    ws2.cell(ws2_line_i, 1).value = ws.cell(ws_line_i, 4).value
                    (coord_e, coord_n) = ws.cell(ws_line_i, 9).value.split(" ")
                    ws2.cell(ws2_line_i, 1).hyperlink = os.environ["FEEDBACK_COMMUNES_URL_CONSULTATION_" + environ + "_ISSUE_22_SITN_COORD"].format(coord_e, coord_n)
                    ws2.cell(ws2_line_i, 1).style = "Hyperlink"
                    ws2.cell(ws2_line_i, 2).value = ws.cell(ws_line_i, 5).value
                    ws2.cell(ws2_line_i, 3).value = ws.cell(ws_line_i, 6).value
                    ws2.cell(ws2_line_i, 4).value = ws.cell(ws_line_i, 7).value
                    ws2.cell(ws2_line_i, 5).value = coord_e
                    ws2.cell(ws2_line_i, 6).value = coord_n
                    ws2.cell(ws2_line_i, 7).value = ws.cell(ws_line_i, 12).value
                    ws2.cell(ws2_line_i, 8).value = _get_issue(ws.cell(ws_line_i, 12).value, issue_solution)
                    ws2_line_i += 1

            ws_line_i += 1

        table_end_row = ws2_line_i - 1

    if table_start_row is not None and table_end_row > table_start_row:
        ref = "A" + str(table_start_row) + ":" + get_column_letter(8) + str(table_end_row)
        _createExcelTable(ws2, "Liste6", ref)

    if table_start_row is None:
        nb_errors_by_list["Liste_6"] = 0
    else:
        nb_errors_by_list["Liste_6"] = table_end_row - table_start_row
    if log:
        print("Liste 6 - end")

    #####################
    #  ISSUE 22
    #####################
    if log:
        print("Issue 22 - start")
    anyI22 = False
    for i22 in issue22_list:
        if i22["COM_FOSNR"] == commune_id:
            if anyI22 is False:
                ws2_line_i += 1
                ws2.cell(ws2_line_i, 1).value = "Bâtiments manquants"
                ws2.cell(ws2_line_i, 1).style = "Headline 1"

                ws2_line_i += 1
                table_start_row = ws2_line_i

                ws2.cell(ws2_line_i, 1).value = "COORDE"
                ws2.cell(ws2_line_i, 2).value = "COORDN"
                ws2.cell(ws2_line_i, 3).value = "LINK"
                ws2.cell(ws2_line_i, 4).value = "DESIGNATION_MO"

                ws2_line_i += 1

                anyI22 = True

            ws2.cell(ws2_line_i, 1).value = i22["BDG_E"]
            ws2.cell(ws2_line_i, 2).value = i22["BDG_N"]
            ws2.cell(ws2_line_i, 3).value = "sitn.ne.ch"
            ws2.cell(ws2_line_i, 3).hyperlink = os.environ["FEEDBACK_COMMUNES_URL_CONSULTATION_" + environ + "_ISSUE_22_SITN_COORD"].format(i22["BDG_E"], i22["BDG_N"])
            ws2.cell(ws2_line_i, 3).style = "Hyperlink"
            ws2.cell(ws2_line_i, 4).value = i22["DESIGNATION_MO"]

            ws2_line_i += 1

    if anyI22 is True:
        table_end_row = ws2_line_i - 1

    if anyI22 is True and table_start_row is not None and table_end_row > table_start_row:
        ref = "A" + str(table_start_row) + ":" + get_column_letter(4) + str(table_end_row)
        _createExcelTable(ws2, "issue22", ref)

    if table_start_row is None:
        nb_errors_by_list["Issue_22"] = 0
    else:
        nb_errors_by_list["Issue_22"] = table_end_row - table_start_row
    if log:
        print("Issue 22 - end")

    # ajuster la largeur des colonnes
    for i in range(10):
        ws2.column_dimensions[get_column_letter(i + 1)].width = "11"

    # supprimer toutes les colonnes autres que "resume"
    for ws_name in wb.sheetnames:
        if not ws_name == "resume":
            wb.remove(wb[ws_name])

    wb.save(feedback_commune_filepath)

    # enregistrer la whiteliste
    if ws_whitelist is not None:
        wb_whitelist.save(whitelist_path)

    return (feedback_commune_filepath, feedback_commune, nb_errors_by_list)


def generateCantonErrorFile(feedback_canton_filepath, issue_solution, today=datetime.strftime(datetime.now(), "%Y%m%d"), environ="INTER", log=False, egidwhitelist=[]):
    # copy canton_file to commune_file
    # feedback_commune_filename = "_".join(["Canton_de_Neuchâtel", "SGRF", "feedback", today]) + ".xlsx"
    feedback_commune_filename = "_".join(["0", "Canton_de_Neuchâtel", "feedback", today]) + ".xlsx"

    feedback_commune_filepath = os.path.join(os.environ["FEEDBACK_COMMUNES_WORKING_DIR"], today, feedback_commune_filename)
    shutil.copy2(feedback_canton_filepath, feedback_commune_filepath)

    nb_errors_by_list = {}
    nb_errors_by_list["Commune_id"] = 0
    nb_errors_by_list["Commune"] = "Canton de Neuchâtel"

    feedback_commune = {"commune_id": 0, "commune_nom": "Canton de Neuchâtel"}

    wb = load_workbook(feedback_commune_filepath)
    if log:
        print(f"${feedback_commune_filepath} loaded")

    # Create resume sheet
    ws2 = wb.create_sheet("resume")

    #####################
    #  INFOS GENERALES
    #####################
    if log:
        print("Infos gen - start")
    ws = wb["Communes"]

    # ws2.cell(1, 1).value = commune_id
    # ws2.cell(1, 1).style = "Title"
    ws2.cell(1, 2).value = "Canton de Neuchâtel"
    ws2.cell(1, 2).style = "Title"
    ws2.cell(1, 4).value = datetime.strftime(datetime.now(), "%d.%m.%Y")
    ws2.cell(1, 6).value = "Explicatif sur la manière de traiter les incohérences"
    ws2.cell(1, 6).hyperlink = "https://www.housing-stat.ch/files/Traitement_erreurs_FR.pdf"
    ws2.cell(1, 6).style = "Hyperlink"

    ws2_line_i = 3
    if log:
        print("Infos gen - end")

    (table_start_row, table_end_row) = (0, 0)

    #####################
    #  LISTE 1
    #####################
    if log:
        print("Liste 1 - start")
    ws = wb["Liste 1"]
    ws_line_i = 7

    table_start_row = None

    if ws_line_i is not None:
        while ws.cell(ws_line_i, 2).value is not None:
            if ws.cell(ws_line_i, 4).value >= 500000000:
                if table_start_row is None:
                    tableHeader = ["COMMUNE", "EGID", "STRNAME", "DEINR", "PLZ4", "PLZNAME"]
                    ws2_line_i, table_start_row = _tableTitleGenerator(ws2, "LISTE 1 - Bâtiments sans coordonnées", tableHeader, ws2_line_i)

                if ws.cell(ws_line_i, 4).value in egidwhitelist.keys():
                    _log_whitelisted_egid(ws.cell(ws_line_i, 4).value, egidwhitelist[ws.cell(ws_line_i, 4).value], today, ws.cell(ws_line_i, 3).value)

                else:
                    ws2.cell(ws2_line_i, 1).value = ws.cell(ws_line_i, 3).value
                    ws2.cell(ws2_line_i, 2).value = ws.cell(ws_line_i, 4).value
                    ws2.cell(ws2_line_i, 3).value = ws.cell(ws_line_i, 11).value
                    ws2.cell(ws2_line_i, 4).value = ws.cell(ws_line_i, 12).value
                    ws2.cell(ws2_line_i, 5).value = ws.cell(ws_line_i, 13).value
                    ws2.cell(ws2_line_i, 6).value = ws.cell(ws_line_i, 15).value
                    ws2_line_i += 1

            ws_line_i += 1

        table_end_row = ws2_line_i - 1

    if table_start_row is not None and table_end_row > table_start_row:
        ref = "A" + str(table_start_row) + ":" + get_column_letter(6) + str(table_end_row)
        _createExcelTable(ws2, "Liste1", ref)

    if table_start_row is None:
        nb_errors_by_list["Liste_1"] = 0
    else:
        nb_errors_by_list["Liste_1"] = table_end_row - table_start_row
    if log:
        print("Liste 1 - end")

    #####################
    #  LISTE 2
    #####################
    if log:
        print("Liste 2 - start")
    ws = wb["Liste 2"]
    ws_line_i = 7

    table_start_row = None

    if ws_line_i is not None:
        while ws.cell(ws_line_i, 2).value is not None:
            if ws.cell(ws_line_i, 4).value and ws.cell(ws_line_i, 4).value >= 500000000:
                if table_start_row is None:
                    tableHeader = ["COMMUNE", "EGID", "Adresse", "GKODE", "ETAT_MO", "GKODN"]
                    ws2_line_i, table_start_row = _tableTitleGenerator(ws2, "LISTE 2 - Coordonnées en dehors de la commune", tableHeader, ws2_line_i)

                if ws.cell(ws_line_i, 4).value in egidwhitelist.keys():
                    _log_whitelisted_egid(ws.cell(ws_line_i, 4).value, egidwhitelist[ws.cell(ws_line_i, 4).value], today, ws.cell(ws_line_i, 3).value)

                else:
                    ws2.cell(ws2_line_i, 1).value = ws.cell(ws_line_i, 3).value
                    ws2.cell(ws2_line_i, 2).value = ws.cell(ws_line_i, 4).value
                    ws2.cell(ws2_line_i, 2).hyperlink = os.environ["FEEDBACK_COMMUNES_URL_CONSULTATION_" + environ + "_ISSUE_22_SITN_COORD"].format(ws.cell(ws_line_i, 11).value, ws.cell(ws_line_i, 12).value)
                    ws2.cell(ws2_line_i, 2).style = "Hyperlink"
                    ws2.cell(ws2_line_i, 3).value = ws.cell(ws_line_i, 5).value
                    ws2.cell(ws2_line_i, 4).value = ws.cell(ws_line_i, 11).value
                    ws2.cell(ws2_line_i, 5).value = ws.cell(ws_line_i, 19).value
                    ws2.cell(ws2_line_i, 6).value = ws.cell(ws_line_i, 12).value
                    ws2_line_i += 1

            ws_line_i += 1

        table_end_row = ws2_line_i - 1

    if table_start_row is not None and table_end_row > table_start_row:
        ref = "A" + str(table_start_row) + ":" + get_column_letter(6) + str(table_end_row)
        _createExcelTable(ws2, "Liste2", ref)

    if table_start_row is None:
        nb_errors_by_list["Liste_2"] = 0
    else:
        nb_errors_by_list["Liste_2"] = table_end_row - table_start_row
    if log:
        print("Liste 2 - end")

    #####################
    #  LISTE 3
    #####################
    if log:
        print("Liste 3 - start")
    ws = wb["Liste 3"]
    ws_line_i = 7

    table_start_row = None

    if ws_line_i is not None:
        while ws.cell(ws_line_i, 2).value is not None:
            if ws.cell(ws_line_i, 4).value >= 500000000:
                if table_start_row is None:
                    tableHeader = ["COMMUNE", "EGID", "PLZ4 RegBL", "PLZ4_Name RegBL", "PLZ4 MO", "ETAT_MO", "PLZ4_Name MO"]
                    ws2_line_i, table_start_row = _tableTitleGenerator(ws2, "LISTE 3 - Divergence de NPA", tableHeader, ws2_line_i)

                if ws.cell(ws_line_i, 4).value in egidwhitelist.keys():
                    _log_whitelisted_egid(ws.cell(ws_line_i, 4).value, egidwhitelist[ws.cell(ws_line_i, 4).value], today, ws.cell(ws_line_i, 3).value)

                else:
                    ws2.cell(ws2_line_i, 1).value = ws.cell(ws_line_i, 3).value
                    ws2.cell(ws2_line_i, 2).value = ws.cell(ws_line_i, 4).value
                    ws2.cell(ws2_line_i, 2).hyperlink = os.environ["FEEDBACK_COMMUNES_URL_CONSULTATION_" + environ + "_ISSUE_22_SITN_COORD"].format(ws.cell(ws_line_i, 20).value, ws.cell(ws_line_i, 21).value)
                    ws2.cell(ws2_line_i, 2).style = "Hyperlink"
                    ws2.cell(ws2_line_i, 3).value = ws.cell(ws_line_i, 8).value
                    ws2.cell(ws2_line_i, 4).value = ws.cell(ws_line_i, 9).value
                    ws2.cell(ws2_line_i, 5).value = ws.cell(ws_line_i, 11).value
                    ws2.cell(ws2_line_i, 6).value = ws.cell(ws_line_i, 27).value
                    ws2.cell(ws2_line_i, 7).value = ws.cell(ws_line_i, 12).value
                    ws2_line_i += 1

            ws_line_i += 1

        table_end_row = ws2_line_i - 1

    if table_start_row is not None and table_end_row > table_start_row:
        ref = "A" + str(table_start_row) + ":" + get_column_letter(7) + str(table_end_row)
        _createExcelTable(ws2, "Liste3", ref)

    if table_start_row is None:
        nb_errors_by_list["Liste_3"] = 0
    else:
        nb_errors_by_list["Liste_3"] = table_end_row - table_start_row
    if log:
        print("Liste 3 - end")

    #####################
    #  LISTE 4
    #####################
    if log:
        print("Liste 4 - start")
    ws = wb["Liste 4"]
    ws_line_i = 7

    table_start_row = None

    if ws_line_i is not None:
        while ws.cell(ws_line_i, 2).value is not None:
            if ws.cell(ws_line_i, 4).value >= 500000000:
                if table_start_row is None:
                    tableHeader = ["COMMUNE", "EGID", "GKAT", "GPARZ", "GEBNR", "STRNAME", "DEINR", "PLZ4", "GBEZ", "ETAT_MO", "BUR / REE"]
                    ws2_line_i, table_start_row = _tableTitleGenerator(ws2, "LISTE 4 - Doublets d'adresses", tableHeader, ws2_line_i)

                if ws.cell(ws_line_i, 4).value in egidwhitelist.keys():
                    _log_whitelisted_egid(ws.cell(ws_line_i, 4).value, egidwhitelist[ws.cell(ws_line_i, 4).value], today, ws.cell(ws_line_i, 3).value)

                else:
                    ws2.cell(ws2_line_i, 1).value = ws.cell(ws_line_i, 3).value
                    ws2.cell(ws2_line_i, 2).value = ws.cell(ws_line_i, 4).value
                    ws2.cell(ws2_line_i, 2).hyperlink = os.environ["FEEDBACK_COMMUNES_URL_CONSULTATION_" + environ + "_ISSUE_22_SITN_COORD"].format(ws.cell(ws_line_i, 7).value, ws.cell(ws_line_i, 8).value)
                    ws2.cell(ws2_line_i, 2).style = "Hyperlink"
                    ws2.cell(ws2_line_i, 3).value = ws.cell(ws_line_i, 6).value
                    ws2.cell(ws2_line_i, 4).value = ws.cell(ws_line_i, 22).value
                    ws2.cell(ws2_line_i, 5).value = ws.cell(ws_line_i, 23).value
                    ws2.cell(ws2_line_i, 6).value = ws.cell(ws_line_i, 11).value
                    ws2.cell(ws2_line_i, 7).value = ws.cell(ws_line_i, 12).value
                    ws2.cell(ws2_line_i, 8).value = ws.cell(ws_line_i, 13).value
                    ws2.cell(ws2_line_i, 9).value = ws.cell(ws_line_i, 14).value
                    ws2.cell(ws2_line_i, 10).value = ws.cell(ws_line_i, 26).value
                    ws2.cell(ws2_line_i, 11).value = ws.cell(ws_line_i, 24).value
                    ws2_line_i += 1

            ws_line_i += 1

        table_end_row = ws2_line_i - 1

    if table_start_row is not None and table_end_row > table_start_row:
        ref = "A" + str(table_start_row) + ":" + get_column_letter(11) + str(table_end_row)
        _createExcelTable(ws2, "Liste4", ref)

    if table_start_row is None:
        nb_errors_by_list["Liste_4"] = 0
    else:
        nb_errors_by_list["Liste_4"] = table_end_row - table_start_row
    if log:
        print("Liste 4 - end")

    #####################
    #  LISTE 5
    #####################
    if log:
        print("Liste 5 - start")
    ws = wb["Liste 5"]
    ws_line_i = 7

    table_start_row = None

    if ws_line_i is not None:
        while ws.cell(ws_line_i, 2).value is not None:
            if ws.cell(ws_line_i, 4).value >= 500000000:
                if table_start_row is None:
                    tableHeader = ["COMMUNE", "EGID", "GKAT", "GKLAS", "GSTAT", "GKODE", "GKODN", "ISSUE", "ETAT_MO", "RESOLUTION_SGRF"]
                    ws2_line_i, table_start_row = _tableTitleGenerator(ws2, "LISTE 5 - Définition du bâtiment", tableHeader, ws2_line_i)

                if ws.cell(ws_line_i, 4).value in egidwhitelist.keys():
                    _log_whitelisted_egid(ws.cell(ws_line_i, 4).value, egidwhitelist[ws.cell(ws_line_i, 4).value], today, ws.cell(ws_line_i, 3).value)

                else:
                    ws2.cell(ws2_line_i, 1).value = ws.cell(ws_line_i, 3).value
                    ws2.cell(ws2_line_i, 2).value = ws.cell(ws_line_i, 4).value
                    (coord_e, coord_n) = ws.cell(ws_line_i, 9).value.split(" ")
                    ws2.cell(ws2_line_i, 2).hyperlink = os.environ["FEEDBACK_COMMUNES_URL_CONSULTATION_" + environ + "_ISSUE_22_SITN_COORD"].format(coord_e, coord_n)
                    ws2.cell(ws2_line_i, 2).style = "Hyperlink"
                    ws2.cell(ws2_line_i, 3).value = ws.cell(ws_line_i, 5).value
                    ws2.cell(ws2_line_i, 4).value = ws.cell(ws_line_i, 6).value
                    ws2.cell(ws2_line_i, 5).value = ws.cell(ws_line_i, 7).value
                    ws2.cell(ws2_line_i, 6).value = coord_e
                    ws2.cell(ws2_line_i, 7).value = coord_n
                    ws2.cell(ws2_line_i, 8).value = ws.cell(ws_line_i, 12).value
                    ws2.cell(ws2_line_i, 9).value = ws.cell(ws_line_i, 14).value
                    ws2.cell(ws2_line_i, 10).value = _get_issue(ws.cell(ws_line_i, 12).value, issue_solution)
                    ws2_line_i += 1

            ws_line_i += 1

        table_end_row = ws2_line_i - 1

    if table_start_row is not None and table_end_row > table_start_row:
        ref = "A" + str(table_start_row) + ":" + get_column_letter(10) + str(table_end_row)
        _createExcelTable(ws2, "Liste5", ref)

    if table_start_row is None:
        nb_errors_by_list["Liste_5"] = 0
    else:
        nb_errors_by_list["Liste_5"] = table_end_row - table_start_row
    if log:
        print("Liste 5 - end")

    #####################
    #  LISTE 6
    #####################
    if log:
        print("Liste 6 - start")
    ws = wb["Liste 6"]
    ws_line_i = 7

    table_start_row = None

    if ws_line_i is not None:
        while ws.cell(ws_line_i, 2).value is not None:
            if ws.cell(ws_line_i, 4).value >= 500000000:
                if table_start_row is None:
                    tableHeader = ["COMMUNE", "EGID", "GKAT", "GKLAS", "GSTAT", "GKODE", "GKODN", "ISSUE", "ETAT_MO", "RESOLUTION_SGRF"]
                    ws2_line_i, table_start_row = _tableTitleGenerator(ws2, "Liste 6 - Catégorie du bâtiment", tableHeader, ws2_line_i)

                if ws.cell(ws_line_i, 4).value in egidwhitelist.keys():
                    _log_whitelisted_egid(ws.cell(ws_line_i, 4).value, egidwhitelist[ws.cell(ws_line_i, 4).value], today, ws.cell(ws_line_i, 3).value)

                else:
                    ws2.cell(ws2_line_i, 1).value = ws.cell(ws_line_i, 3).value
                    ws2.cell(ws2_line_i, 2).value = ws.cell(ws_line_i, 4).value
                    (coord_e, coord_n) = ws.cell(ws_line_i, 9).value.split(" ")
                    ws2.cell(ws2_line_i, 2).hyperlink = os.environ["FEEDBACK_COMMUNES_URL_CONSULTATION_" + environ + "_ISSUE_22_SITN_COORD"].format(coord_e, coord_n)
                    ws2.cell(ws2_line_i, 2).style = "Hyperlink"
                    ws2.cell(ws2_line_i, 3).value = ws.cell(ws_line_i, 5).value
                    ws2.cell(ws2_line_i, 4).value = ws.cell(ws_line_i, 6).value
                    ws2.cell(ws2_line_i, 5).value = ws.cell(ws_line_i, 7).value
                    ws2.cell(ws2_line_i, 6).value = coord_e
                    ws2.cell(ws2_line_i, 7).value = coord_n
                    ws2.cell(ws2_line_i, 8).value = ws.cell(ws_line_i, 12).value
                    ws2.cell(ws2_line_i, 9).value = ws.cell(ws_line_i, 14).value
                    ws2.cell(ws2_line_i, 10).value = _get_issue(ws.cell(ws_line_i, 12).value, issue_solution)
                    ws2_line_i += 1

            ws_line_i += 1

        table_end_row = ws2_line_i - 1

    if table_start_row is not None and table_end_row > table_start_row:
        ref = "A" + str(table_start_row) + ":" + get_column_letter(10) + str(table_end_row)
        _createExcelTable(ws2, "Liste6", ref)

    if table_start_row is None:
        nb_errors_by_list["Liste_6"] = 0
    else:
        nb_errors_by_list["Liste_6"] = table_end_row - table_start_row
    if log:
        print("Liste 6 - end")

    #####################
    #  ISSUE 22
    #####################
    if log:
        print("Issue 22 - start")

    nb_errors_by_list["Issue_22"] = 0
    if log:
        print("Issue 22 - end")

    # ajuster la largeur des colonnes
    for i in range(10):
        ws2.column_dimensions[get_column_letter(i + 1)].width = "11"

    # supprimer toutes les colonnes autres que "resume"
    for ws_name in wb.sheetnames:
        if not ws_name == "resume":
            wb.remove(wb[ws_name])

    wb.save(feedback_commune_filepath)

    return (feedback_commune_filepath, feedback_commune, nb_errors_by_list)


def createDBSession(env_substring):
    engine = create_engine("postgresql+psycopg2://{}:{}@{}:{}/{}".format(os.environ[env_substring + "_USERNAME"], os.environ[env_substring + "_PASSWORD"], os.environ[env_substring + "_HOST"], os.environ[env_substring + "_PORT"], os.environ[env_substring + "_DATABASE"]))

    # create session
    Session = sessionmaker()
    Session.configure(bind=engine)
    return Session()
