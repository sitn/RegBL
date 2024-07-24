import shutil
from openpyxl import load_workbook, utils
import re
from dotenv import load_dotenv
import os

load_dotenv(override=True)


# Chemin du fichier Excel original
original_file_path = os.getenv("I22_ORIGINAL_FILE_PATH")  # Remplacez par le chemin de votre fichier
sheet_name = "NoCommuneOFS"  # Nom de la feuille contenant les données
output_path = os.getenv("I22_DESTINATION_PATH")


# Fonction pour remplacer les références de lignes dans les formules
def _remplacer_references(formule, ligne):
    # Expression régulière pour trouver des références de cellules (par exemple B1, C5)
    pattern = re.compile(r"([A-Z]+)(\d+)")

    def remplacement(match):
        col, _ = match.groups()
        return f"{col}{ligne}"

    # Remplacer toutes les occurrences des références de lignes par la ligne courante
    return pattern.sub(remplacement, formule)


def _get_table(feuille, cell_reference="A1"):
    # Obtenir la coordonnée de la cellule sous forme d'objet Cell
    cell = feuille[cell_reference]

    # Parcourir tous les tableaux de la feuille de calcul
    for tableau in feuille.tables.values():
        # print("tableau:", tableau)
        # print(f"utils.cell.range_boundaries({tableau.ref})", utils.cell.range_boundaries(tableau.ref))
        # Obtenir les limites du tableau
        min_col, min_row, max_col, max_row = utils.cell.range_boundaries(tableau.ref)

        # Vérifier si la cellule se trouve dans les limites du tableau
        if min_row <= cell.row <= max_row and min_col <= cell.column <= max_col:
            return tableau

    return None


# Charger le fichier original avec openpyxl
wb_original = load_workbook(original_file_path)
ws_original = wb_original[sheet_name]

# Lire les numéros de commune uniques (en supposant qu'ils sont dans la colonne D)
communes = set()

for row in ws_original.iter_rows(min_row=2, min_col=4, max_col=4, values_only=True):
    communes.add(row[0])

# Créer une copie du fichier original pour chaque numéro de commune
for commune in communes:
    # Définir le chemin du nouveau fichier
    new_file_path = f"{output_path}{commune}.xlsx"

    # Copier le fichier original
    shutil.copyfile(original_file_path, new_file_path)

    # Charger le fichier copié avec openpyxl
    wb_new = load_workbook(new_file_path)
    ws_new = wb_new[sheet_name]

    # Lire toutes les lignes et filtrer celles qui ne correspondent pas à la commune
    rows_to_delete = []

    for row in ws_new.iter_rows(min_row=2, values_only=False):
        if row[3].value != commune:
            rows_to_delete.append(row[0].row)

    # Supprimer les lignes inutiles en commençant par la fin pour éviter les décalages
    for row in sorted(rows_to_delete, reverse=True):
        ws_new.delete_rows(row, 1)

    tbl = _get_table(ws_new)
    tbl.ref = f"A1:AH{ws_new.max_row}"

    cols2update = ["N", "P", "Q", "AF"]
    for row in range(1, ws_new.max_row + 1):
        for col in cols2update:
            cellule = ws_new[f"{col}{row}"]
            if cellule.value and isinstance(cellule.value, str) and cellule.value.startswith("="):
                # print('cellule.value:', cellule.value)
                value = _remplacer_references(cellule.value, row)
                # print('replaced by:  ', value)
                # print('')
                cellule.value = value

    # print("commune", commune)
    ws_new.title = str(commune)

    # Sauvegarder le nouveau fichier Excel
    wb_new.save(new_file_path)


print("La création des fichiers Excel est terminée.")
