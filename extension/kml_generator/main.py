from pyproj import CRS, Transformer
import simplekml
import openpyxl
import re
import os

from dotenv import load_dotenv

load_dotenv(r'..\..\.env')

input_path = os.environ['KML_GENERATOR_INPUT_PATH']
output_path = os.environ['KML_GENERATOR_OUTPUT_PATH']


crs_2056 = CRS.from_epsg(2056)
crs_4326 = CRS.from_epsg(4326)
TF = Transformer.from_crs(crs_2056, crs_4326, always_xy=True)


def xlsx2kml(filepath, sheetname, kml_filename):
    kml = simplekml.Kml()
    
    wb=openpyxl.load_workbook(filepath, data_only=True)
    ws = wb[sheetname]
    
    i=1
    while i<10000:
        i += 1
        
        if ws.cell(i, 1).value is None:
            break

        if ws.cell(i,2).value != 'MO slt sans adresse' and ws.cell(i,2).value != 'MO slt avec adresse':
            continue

        # adresse = ws.cell(i,19).value or ws.cell(i,8).value + " " + str(ws.cell(i,20).value or ws.cell(i,9).value) + ", " + str(ws.cell(i,10).value) + " " + ws.cell(i,11).value
        # pnt = kml.newpoint(
        #     name = ', '.join(filter(None, [ws.cell(i,1).value, adresse])),
        #     coords = [TF.transform(float(ws.cell(i,30).value), float(ws.cell(i,31).value))],
        # )

        # if ws.cell(i,21).value == 'oui':
        #     # delete = oui
        #     continue

        # if ws.cell(i, 50).value is not None:
        # # if ws.cell(i, 50).value is None:
        #     # Cellule pour les zones des communes (si elles existes, par exemple pour les zones littoral)
        #     continue

        # pnt = kml.newpoint(
        #     name = ws.cell(i,1).value,
        #     # name = ', '.join(filter(None, [ws.cell(i,1).value, '/'.join(filter(None, [str(ws.cell(i,12).value), str(ws.cell(i,13).value)]))])),
        #     coords = [TF.transform(float(ws.cell(i,30).value), float(ws.cell(i,31).value))],
        # )
        # pnt.style.labelstyle.color = simplekml.Color.green

        # ####################################################
        # #   AFFICHE LES PROBLEMES DANS LE CTRL3 EN ROUGE   #
        # ####################################################
        # if (ws.cell(i,14).value == "Problem") and (ws.cell(i,12).value == 1080): # Ctrl 1 = "Problem" and GKAT = 1080
        #     pnt = kml.newpoint(
        #         name = ', '.join(filter(None, [ws.cell(i,1).value, ws.cell(i,13).value])),
        #         coords = [TF.transform(float(ws.cell(i,30).value), float(ws.cell(i,31).value))],
        #     )
        #     pnt.style.labelstyle.color = simplekml.Color.red
        
        ############################################
        #   AFFICHE LES ADRESSES DANS LE EN VERT   #
        ############################################
        
        if ws.cell(i,21).value == 'oui':
            continue

        pnt = kml.newpoint(
            name = ', '.join(filter(None, [ws.cell(i,1).value, ws.cell(i,39).value])),
            coords = [TF.transform(float(ws.cell(i,30).value), float(ws.cell(i,31).value))],
        )
        pnt.style.labelstyle.color = simplekml.Color.green
        
        ####################################################
        #   AFFICHE LES PROBLEMES DANS LE CTRL3 EN ROUGE   #
        ####################################################
        # if (ws.cell(i,32).value == "Problem"): # Ctrl 3 = "Problem"
        #     pnt.style.labelstyle.color = simplekml.Color.red
        
        
        ##########################################
        #   AFFICHE LES DELETE == OUI EN ROUGE   #
        ##########################################
        # if (ws.cell(i,21).value == 'oui'): # delete = oui
        #     pnt.style.labelstyle.color = simplekml.Color.red
        
        
        ##########################################################
        #   N'AFFICHE QUE LES PROBLEMES DANSL E CTRL3 EN ROUGE   #
        ##########################################################
        # if (ws.cell(i,32).value == "Problem"): # Control3 = "Problem"
        #     pnt = kml.newpoint(
        #         name = ', '.join(filter(None, [ws.cell(i,1).value, ws.cell(i,39).value])),
        #         coords = [TF.transform(float(ws.cell(i,30).value), float(ws.cell(i,31).value))],
        #     )
        #     pnt.style.labelstyle.color = simplekml.Color.red
        
        
        #############################################################
        #   N'AFFICHE QUE LES DELETE = OUI DANSL E CTRL3 EN ROUGE   #
        #############################################################
        # if (ws.cell(i,21).value == 'oui'): # delete = oui
        #     pnt = kml.newpoint(
        #         name = ', '.join(filter(None, [ws.cell(i,1).value, str(ws.cell(i,49).value)])),
        #         coords = [TF.transform(float(ws.cell(i,30).value), float(ws.cell(i,31).value))],
        #     )
        #     pnt.style.labelstyle.color = simplekml.Color.red

    kml.save(kml_filename)
    return




if __name__ == "__main__":
    files = os.listdir(input_path)
    for filename in files:
        if re.match(r'6[4,5]{1}[\d]{2}\.xlsx', filename) is not None:
            no_commune = filename.split('.')[0]
            print(no_commune)
            input_filepath = os.path.join(input_path, filename)
            kml_filename = no_commune + ".kml"
            output_filepath = os.path.join(output_path, kml_filename)
            xlsx2kml(input_filepath, no_commune, output_filepath)
