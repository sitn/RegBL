from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from datetime import datetime
import os
from dotenv import load_dotenv

load_dotenv()


working_dir = os.getenv("PERIODE_CONSTRUCTION_ANALYZER_DIRECTORY")


if __name__ == "__main__":

    # create workbook excel
    wb = Workbook()
    ws = wb.active

    # titles
    i = 1  # line index
    ws.cell(i, 1).value = "TUILE"
    ws.cell(i, 2).value = "EGID"
    ws.cell(i, 3).value = "LAST_YEAR_WITHOUT"
    ws.cell(i, 4).value = "FIRST_YEAR_WITH"
    ws.cell(i, 5).value = "REMARK"
    ws.cell(i, 6).value = "Lien_image"

    for root, dirs, files in os.walk(working_dir):
        tile = root.rsplit(os.path.sep, 1)[1]

        for f in files:
            if f.endswith(".csv"):
                input_csv = os.path.join(root, f)
                print(input_csv)

                with open(input_csv, "r") as file:
                    title_row = True
                    for x in file:
                        if title_row is True:
                            title_row = False
                            continue

                        if len(x) > 0:
                            i += 1
                            for j, y in enumerate(x.split(",")):
                                ws.cell(i, 1).value = tile
                                ws.cell(i, j + 2).value = y.replace("\n", "")

                            image_path = os.path.join(working_dir, tile, x.split(",")[0] + ".png")
                            if os.path.exists(image_path):
                                ws.cell(i, j + 3).hyperlink = image_path
                                ws.cell(i, j + 3).value = "image"
                                ws.cell(i, j + 3).style = "Hyperlink"

    # Generate table
    tab = Table(displayName="Table1", ref="A1:F{}".format(i))

    # Add a default style with striped rows and banded columns
    style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    tab.tableStyleInfo = style

    ws.add_table(tab)
    # ws.auto_filter.ref = ws.dimensions

    wb.save(os.path.join(working_dir, datetime.strftime(datetime.now(), "%Y%m%d-%H%M%S") + "_merge_analyzer.xlsx"))
