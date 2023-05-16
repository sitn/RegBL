from openpyxl import load_workbook
import os
import sys
from dotenv import load_dotenv
from models import RegBLApurementFeedbackHebdoCommunes

load_dotenv(r'..\..\.env')

sys.path.insert(0, r'..\..\utils')
import utils


session = utils.createDBSession('ARCHIVE_FEEDBACK_COMMUNES_DB')

# lire le fichier excel
source_path = os.environ['ARCHIVE_FEEDBACK_COMMUNES_EXCEL_PATH']
files = os.listdir(source_path)
for f in files:
    if not f.endswith('.xlsx'):
        continue

    file_path = os.path.join(source_path, f)

    wb = load_workbook(file_path)
    ws = wb['Communes']


    # get date and check if it already exists in database (if yes: skip this file)
    date = ws.cell(1,2).value.replace('Etat: ', '').split('.')
    date = '-'.join(date[::-1]) if len(date) == 3 else None

    test = session.query(
        RegBLApurementFeedbackHebdoCommunes
    ).filter(
        RegBLApurementFeedbackHebdoCommunes.date_version == date
    ).all()

    if len(test) > 0:
        continue


    line_i = utils._findRowIndex(ws, 'Canton', column_id=2)
    line_i += 2
    
    while ws.cell(line_i, 2).value == 'NE':
        data = None
        data = RegBLApurementFeedbackHebdoCommunes()
        data.no_commune_ofs = ws.cell(line_i, 3).value
        data.commune = ws.cell(line_i, 4).value
        data.batiments = ws.cell(line_i, 5).value
        data.entrees = ws.cell(line_i, 6).value
        data.batiments_manquants = ws.cell(line_i, 9).value
        data.liste_1 = ws.cell(line_i, 12).value
        data.liste_1_pc = ws.cell(line_i, 14).value
        data.liste_2 = ws.cell(line_i, 17).value
        data.liste_2_pc = ws.cell(line_i, 19).value
        data.liste_3 = ws.cell(line_i, 22).value
        data.liste_3_pc = ws.cell(line_i, 24).value
        data.liste_4 = ws.cell(line_i, 27).value
        data.liste_4_pc = ws.cell(line_i, 29).value
        data.liste_5 = ws.cell(line_i, 32).value
        data.liste_5_pc = ws.cell(line_i, 34).value
        data.liste_6 = ws.cell(line_i, 37).value
        data.liste_6_pc = ws.cell(line_i, 39).value
        data.total_listes_pc = ws.cell(line_i, 41).value
        data.ext_batiments = ws.cell(line_i, 42).value
        data.ext_batiments_gklas = ws.cell(line_i, 43).value
        data.ext_batiments_gklas_pc = ws.cell(line_i, 44).value
        data.ext_batiments_gbaup = ws.cell(line_i, 45).value
        data.ext_batiments_gbaup_pc = ws.cell(line_i, 46).value
        data.ext_batiments_surf30_batiments = ws.cell(line_i, 49).value
        data.ext_batiments_surf30_gklas = ws.cell(line_i, 50).value
        data.ext_batiments_surf30_gklas_pc = ws.cell(line_i, 51).value
        data.ext_batiments_surf30_gbaup = ws.cell(line_i, 52).value
        data.ext_batiments_surf30_gbaup_pc = ws.cell(line_i, 53).value
        data.date_version = date

        session.add(data)
        line_i += 1


session.commit()
