# -*- coding: utf-8 -*-
from jinja2 import Environment, FileSystemLoader
import csv
import openpyxl
import smtplib
from email.message import EmailMessage
from email.utils import formatdate
from email.mime.application import MIMEApplication
import os, re, sys
from datetime import datetime

from dotenv import load_dotenv

load_dotenv(override=True)


dt = datetime.strftime(datetime.now(), "%Y%m%d")


def get_mail_template():
    file_loader = FileSystemLoader(r"apurement\mail_communes\templates")
    env = Environment(loader=file_loader)
    return env.get_template("feedback_commune.html")


def get_email_registry():
    wb = openpyxl.load_workbook(r"apurement\mail_communes\static\RegBL_contactCommune.xlsx")
    ws = wb["Communes"]
    row = 2
    mail = {}
    while ws.cell(row, 1).value != None:
        mail[str(ws.cell(row, 2).value)] = ws.cell(row, 6).value
        row += 1

    return mail


def send_mail(to, subject, content, files=[]):
    assert isinstance(to, list)
    assert isinstance(files, list)
    to = [value for value in to if value != None]
    to = [value for value in to if value != ""]

    if len(to) == 0:
        return

    msg = EmailMessage()
    msg["From"] = os.getenv("MAIL_COMMUNE_FROM")
    msg["To"] = ", ".join(to)
    msg["Date"] = formatdate(localtime=True)
    msg["Subject"] = subject
    msg["Body"] = ""
    msg.add_alternative(content, subtype="html")

    for f in files:
        filename = os.path.basename(f)
        with open(f, "rb") as f:
            part = MIMEApplication(f.read(), Name=filename)
        # After the file is closed
        part["Content-Disposition"] = f'attachment; filename="{filename}"'
        msg.attach(part)

    s = smtplib.SMTP(os.getenv("MAIL_SMTP"))
    s.send_message(msg)
    s.quit()
    return


if __name__ == "__main__":
    # email de test
    testemail = None
    if "--testemail" in sys.argv:
        testemail = sys.argv[sys.argv.index("--testemail") + 1]

    feedback_path = os.getenv("MAIL_COMMUNE_FEEDBACK_PATH")
    feedback_csv_path = os.path.join(feedback_path, dt, f"{dt}-feedback.csv")

    jinja_tpl = get_mail_template()
    mails = get_email_registry()

    with open(feedback_csv_path, mode="r", newline="") as file:
        spamreader = csv.reader(file, delimiter=";")

        # parcourir les communes
        for row in spamreader:
            # ne pas tenir compte de la ligne des titres
            if row[0].isnumeric():
                total_error = sum([int(a) for a in row[2:]])

                data_tpl = {
                    "COMMUNE_NAME": row[1],
                    "NB_ERROR_LIST_1": row[2],
                    "NB_ERROR_LIST_2": row[3],
                    "NB_ERROR_LIST_3": row[4],
                    "NB_ERROR_LIST_4": row[5],
                    "NB_ERROR_LIST_5": row[6],
                    "NB_ERROR_LIST_6": row[7],
                    "NB_ERROR_ISSUE_22": row[8],
                    "ADMINISTRATOR_NAME": os.getenv("MAIL_COMMUNE_ADMINISTRATOR_NAME"),
                    "ADMINISTRATOR_EMAIL": os.getenv("MAIL_COMMUNE_ADMINISTRATOR_EMAIL"),
                }

                # remplir le template
                email_content = jinja_tpl.render(data_tpl)

                commune_id = row[0]
                commune_name = row[1].replace(" ", "_")

                attached_file = os.path.join(feedback_path, dt, f"{commune_id}_{commune_name}_feedback_{dt}.xlsx")
                if not os.path.exists:
                    attached_file = []

                # destinataire
                if testemail is not None:
                    if ";" in testemail or "," in testemail:
                        to = re.split(";|,", testemail)
                    else:
                        to = [testemail]
                else:
                    if int(commune_id) > 0:
                        to = [mails[row[0]]]
                        # continue
                    else:
                        # Canton de neuchâtel
                        to = [os.getenv("MAIL_ME")]

                # envoyer le mail avec la pièce jointe si nécessaire
                attached_files = []
                if total_error > 0:
                    attached_files = [attached_file]
                send_mail(to, "RegBL - apurement des données", email_content, files=attached_files)

                mail_filename = f"{row[0]}_{row[1]}_{dt}_mail.html"
                mail_path = os.path.join(os.environ["FEEDBACK_COMMUNES_WORKING_DIR"], dt, mail_filename)
                with open(mail_path, "w", encoding="utf-8") as mail_file:
                    mail_file.write(email_content)
